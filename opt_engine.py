#!/usr/bin/env python3
"""
Motor de optimizacion de rutas - Exportadora de Banano
Yuber: precio base 20P, +$25.000 por pallet adicional hasta 24P.
Viaje combinado Chigorodo+Apartado: +$100.000 entrada Apartado.
Demetrio: puede hacer 2 viajes por dia (Viaje 1 y Viaje 2).
Edwin Echavarria: Mula 24P — $1.050.000 desde Chigorodo, $600.000 desde Apartado.
"""

import pandas as pd
import re
import math
import io
import copy
from itertools import combinations
from openpyxl import Workbook
from openpyxl.styles import Font, PatternFill, Alignment, Border, Side
from openpyxl.utils import get_column_letter

# ── Flota ────────────────────────────────────────────────────
VEHICLE_DISPLAY = {
    'YUBER_MULA_1':        {'conductor': 'Yuber',   'vehicle': 'Mula 1 - 24P'},
    'YUBER_MULA_2':        {'conductor': 'Yuber',   'vehicle': 'Mula 2 - 24P'},
    'YUBER_MULA_3':        {'conductor': 'Yuber',   'vehicle': 'Mula 3 - 24P'},
    'YUBER_MULA_4':        {'conductor': 'Yuber',   'vehicle': 'Mula 4 - 24P'},
    'YUBER_MULA_5':        {'conductor': 'Yuber',   'vehicle': 'Mula 5 - 24P'},
    'YUBER_MULA_6':        {'conductor': 'Yuber',   'vehicle': 'Mula 6 - 24P'},
    'DEMETRIO_PATINETA':   {'conductor': 'Demetrio', 'vehicle': 'Patineta 18P (Viaje 1)'},
    'DEMETRIO_PATINETA_2': {'conductor': 'Demetrio', 'vehicle': 'Patineta 18P (Viaje 2)'},
    'EDWIN_MULA':          {'conductor': 'Edwin',   'vehicle': 'Mula 24P (Viaje 1)'},
    'EDWIN_MULA_2':        {'conductor': 'Edwin',   'vehicle': 'Mula 24P (Viaje 2)'},
}

ALL_VEHICLE_IDS = list(VEHICLE_DISPLAY.keys())

# Yuber: costo = precio base (≤20P). Pallets 21-24: +$25.000 c/u.
_YUBER_CHIGORODO = [
    {'conductor': 'YUBER', 'tipo': 'MULA', 'capacidad': 24, 'costo': 1050000,
     'pallets_negociados': 20, 'costo_extra_pallet': 25000, 'vehicle_id': 'YUBER_MULA_1'},
    {'conductor': 'YUBER', 'tipo': 'MULA', 'capacidad': 24, 'costo': 1050000,
     'pallets_negociados': 20, 'costo_extra_pallet': 25000, 'vehicle_id': 'YUBER_MULA_2'},
    {'conductor': 'YUBER', 'tipo': 'MULA', 'capacidad': 24, 'costo': 1050000,
     'pallets_negociados': 20, 'costo_extra_pallet': 25000, 'vehicle_id': 'YUBER_MULA_3'},
    {'conductor': 'YUBER', 'tipo': 'MULA', 'capacidad': 24, 'costo': 1050000,
     'pallets_negociados': 20, 'costo_extra_pallet': 25000, 'vehicle_id': 'YUBER_MULA_4'},
    {'conductor': 'YUBER', 'tipo': 'MULA', 'capacidad': 24, 'costo': 1050000,
     'pallets_negociados': 20, 'costo_extra_pallet': 25000, 'vehicle_id': 'YUBER_MULA_5'},
    {'conductor': 'YUBER', 'tipo': 'MULA', 'capacidad': 24, 'costo': 1050000,
     'pallets_negociados': 20, 'costo_extra_pallet': 25000, 'vehicle_id': 'YUBER_MULA_6'},
]
_YUBER_APARTADO = [
    {'conductor': 'YUBER', 'tipo': 'MULA', 'capacidad': 24, 'costo': 630000,
     'pallets_negociados': 20, 'costo_extra_pallet': 25000, 'vehicle_id': 'YUBER_MULA_1'},
    {'conductor': 'YUBER', 'tipo': 'MULA', 'capacidad': 24, 'costo': 630000,
     'pallets_negociados': 20, 'costo_extra_pallet': 25000, 'vehicle_id': 'YUBER_MULA_2'},
    {'conductor': 'YUBER', 'tipo': 'MULA', 'capacidad': 24, 'costo': 630000,
     'pallets_negociados': 20, 'costo_extra_pallet': 25000, 'vehicle_id': 'YUBER_MULA_3'},
    {'conductor': 'YUBER', 'tipo': 'MULA', 'capacidad': 24, 'costo': 630000,
     'pallets_negociados': 20, 'costo_extra_pallet': 25000, 'vehicle_id': 'YUBER_MULA_4'},
    {'conductor': 'YUBER', 'tipo': 'MULA', 'capacidad': 24, 'costo': 630000,
     'pallets_negociados': 20, 'costo_extra_pallet': 25000, 'vehicle_id': 'YUBER_MULA_5'},
    {'conductor': 'YUBER', 'tipo': 'MULA', 'capacidad': 24, 'costo': 630000,
     'pallets_negociados': 20, 'costo_extra_pallet': 25000, 'vehicle_id': 'YUBER_MULA_6'},
]

_EDWIN_CHIGORODO = [
    {'conductor': 'EDWIN', 'tipo': 'MULA', 'capacidad': 24, 'costo': 1050000,
     'vehicle_id': 'EDWIN_MULA'},
    {'conductor': 'EDWIN', 'tipo': 'MULA', 'capacidad': 24, 'costo': 1050000,
     'vehicle_id': 'EDWIN_MULA_2'},
]
_EDWIN_APARTADO = [
    {'conductor': 'EDWIN', 'tipo': 'MULA', 'capacidad': 24, 'costo': 600000,
     'vehicle_id': 'EDWIN_MULA'},
    {'conductor': 'EDWIN', 'tipo': 'MULA', 'capacidad': 24, 'costo': 600000,
     'vehicle_id': 'EDWIN_MULA_2'},
]

VEHICLES_BY_ROUTE = {
    ('CHIGORODO', 'PUERTO ANTIOQUIA'): _YUBER_CHIGORODO + _EDWIN_CHIGORODO + [
        {'conductor': 'DEMETRIO', 'tipo': 'PATINETA', 'capacidad': 18, 'costo': 850000,
         'vehicle_id': 'DEMETRIO_PATINETA'},
        {'conductor': 'DEMETRIO', 'tipo': 'PATINETA', 'capacidad': 18, 'costo': 850000,
         'vehicle_id': 'DEMETRIO_PATINETA_2'},
    ],
    ('CHIGORODO', 'UNIBAN ZUNGO'): _YUBER_CHIGORODO + _EDWIN_CHIGORODO + [
        {'conductor': 'DEMETRIO', 'tipo': 'PATINETA', 'capacidad': 18, 'costo': 850000,
         'vehicle_id': 'DEMETRIO_PATINETA'},
        {'conductor': 'DEMETRIO', 'tipo': 'PATINETA', 'capacidad': 18, 'costo': 850000,
         'vehicle_id': 'DEMETRIO_PATINETA_2'},
    ],
    ('APARTADO', 'PUERTO ANTIOQUIA'): _YUBER_APARTADO + _EDWIN_APARTADO + [
        {'conductor': 'DEMETRIO', 'tipo': 'PATINETA', 'capacidad': 18, 'costo': 550000,
         'vehicle_id': 'DEMETRIO_PATINETA'},
        {'conductor': 'DEMETRIO', 'tipo': 'PATINETA', 'capacidad': 18, 'costo': 550000,
         'vehicle_id': 'DEMETRIO_PATINETA_2'},
    ],
    ('APARTADO', 'UNIBAN ZUNGO'): _YUBER_APARTADO + _EDWIN_APARTADO + [
        {'conductor': 'DEMETRIO', 'tipo': 'PATINETA', 'capacidad': 18, 'costo': 550000,
         'vehicle_id': 'DEMETRIO_PATINETA'},
        {'conductor': 'DEMETRIO', 'tipo': 'PATINETA', 'capacidad': 18, 'costo': 550000,
         'vehicle_id': 'DEMETRIO_PATINETA_2'},
    ],
}

CONSOLIDACION_ROUTES = {}

FARM_ZONES = {
    # Apartadó: Doña Francia, Chispero, Santa María, Salvamento
    'SANTA MARIA DEL MONTE': 'APARTADO',
    'STA MARIA DEL MONTE':   'APARTADO',
    'DONA FRANCIA':          'APARTADO',
    'DOÑA FRANCIA':          'APARTADO',
    'CHISPERO':              'APARTADO',
    'SALVAMENTO':            'APARTADO',
    # Chigorodó: San Bartolo, Juana Pío
    'JUANA PIO':             'CHIGORODO',
    'SAN BARTOLO':           'CHIGORODO',
}

# Capacidad máxima disponible al mediodía por finca (pallets).
# Fincas no listadas: sin restricción (salen todos al mediodía si hace falta).
FARM_MEDIODIA_MAX = {
    'CHISPERO':              12,
    'SANTA MARIA DEL MONTE': 16,
    'STA MARIA DEL MONTE':   16,
    'DONA FRANCIA':          20,
    'DOÑA FRANCIA':          20,
    'SAN BARTOLO':           15,
    'JUANA PIO':             13,
    'SALVAMENTO':             6,
}

# Restricciones conductor → fincas que NO puede visitar
# Edwin: su Mula no cabe en las vías de San Bartolo (Chigorodó) ni Santa María (Apartadó)
CONDUCTOR_FARM_RESTRICTIONS = {
    'EDWIN': {'SAN BARTOLO', 'STA MARIA DEL MONTE', 'SANTA MARIA DEL MONTE'},
}

# Mínimos de pallets para cuarteo (viaje con más de una finca)
# Si el viaje ya lleva otra finca y se quiere agregar ésta, debe llegar al mínimo
# Regla: Santa María ≥ 8P, demás fincas ≥ 12P
CUARTEO_MIN_PALLETS = {
    'STA MARIA DEL MONTE':   8,   # vía más complicada, mínimo menor
    'SANTA MARIA DEL MONTE': 8,   # alias sin abreviar
}
CUARTEO_MIN_DEFAULT = 12          # cualquier otra finca en cuarteo

DAY_ORDER  = ['LUNES', 'MARTES', 'MIERCOLES', 'JUEVES', 'VIERNES', 'SABADO']
DAY_EMOJIS = {'LUNES': '🟢', 'MARTES': '🔵', 'MIERCOLES': '🟡', 'JUEVES': '🟠', 'VIERNES': '🔴', 'SABADO': '⚪'}
DAY_COLORS = {'LUNES': '1B5E20', 'MARTES': '0D47A1', 'MIERCOLES': 'F57F17', 'JUEVES': 'E65100', 'VIERNES': 'B71C1C', 'SABADO': '424242'}

ALL_CONDUCTORS = list({v['conductor'] for v in VEHICLE_DISPLAY.values()})


# ── Parser ────────────────────────────────────────────────────
def parse_orders(pedidos_file):
    df = pd.read_excel(pedidos_file, sheet_name=0, header=None)
    day_row = df.iloc[2]
    day_start_cols = {}
    for col_idx, val in enumerate(day_row):
        if pd.notna(val):
            val_upper = str(val).strip().upper()
            for d in DAY_ORDER:
                if val_upper == d:
                    day_start_cols[col_idx] = d
                    break
            if col_idx not in day_start_cols:
                for d in DAY_ORDER:
                    if d in val_upper or val_upper in d:
                        day_start_cols[col_idx] = d
                        break
    if not day_start_cols:
        raise ValueError("No se encontraron dias en fila 3 del Excel")
    sorted_starts = sorted(day_start_cols.keys())
    day_ranges = {}
    for i, start in enumerate(sorted_starts):
        end = sorted_starts[i + 1] if i + 1 < len(sorted_starts) else df.shape[1]
        day_ranges[day_start_cols[start]] = (start, end)
    port_row        = df.iloc[3]
    pallet_size_row = df.iloc[5]
    farm_data       = df.iloc[6:].copy()
    orders = {}
    for day_name, (start_col, end_col) in day_ranges.items():
        port = 'PUERTO ANTIOQUIA'
        if pd.notna(port_row.iloc[start_col]):
            pv = str(port_row.iloc[start_col]).upper()
            if 'ZUNGO' in pv or 'UNIBAN' in pv:
                port = 'UNIBAN ZUNGO'
        pallet_sizes = []
        last_ps = 55
        for col_idx in range(start_col, end_col):
            v = pallet_size_row.iloc[col_idx]
            if pd.notna(v):
                m = re.search(r'(\d+)', str(v))
                if m:
                    last_ps = int(m.group(1))
            pallet_sizes.append(last_ps)
        for _, row in farm_data.iterrows():
            farm_raw = row.iloc[0]
            if pd.isna(farm_raw):
                continue
            farm_name     = str(farm_raw).strip().upper()
            total_cajas   = 0.0
            total_pallets = 0
            pallets_by_size = {}
            for i, col_idx in enumerate(range(start_col, end_col)):
                val = row.iloc[col_idx]
                if pd.isna(val):
                    continue
                try:
                    cajas = float(val)
                except (ValueError, TypeError):
                    continue
                if cajas > 0:
                    ps = pallet_sizes[i]
                    p  = math.ceil(cajas / ps)
                    total_cajas   += cajas
                    total_pallets += p
                    pallets_by_size[ps] = pallets_by_size.get(ps, 0) + p
            if total_cajas > 0:
                if day_name not in orders:
                    orders[day_name] = {}
                if farm_name not in orders[day_name]:
                    orders[day_name][farm_name] = {}
                orders[day_name][farm_name][port] = {
                    'cajas':           int(round(total_cajas)),
                    'pallets':         total_pallets,
                    'pallets_by_size': pallets_by_size,
                }
    return orders


# ── Knapsack 0/1 con snapshot traceback ──────────────────────
def min_cost_assignment_bounded(total_pallets, vehicles):
    P = math.ceil(total_pallets)
    if P <= 0:
        return 0, []
    if not vehicles:
        return float('inf'), []
    INF     = float('inf')
    max_cap = sum(v['capacidad'] for v in vehicles)
    dp_cost = [INF] * (max_cap + 1)
    dp_cost[0] = 0
    snapshots = [dp_cost[:]]
    for v in vehicles:
        cap  = v['capacidad']
        cost = v['costo']
        for j in range(max_cap, cap - 1, -1):
            prev_j = j - cap
            if dp_cost[prev_j] == INF:
                continue
            new_cost = dp_cost[prev_j] + cost
            if new_cost < dp_cost[j]:
                dp_cost[j] = new_cost
        snapshots.append(dp_cost[:])
    best_j = -1
    for j in range(P, max_cap + 1):
        if dp_cost[j] == INF:
            continue
        if best_j == -1 or dp_cost[j] < dp_cost[best_j]:
            best_j = j
        elif dp_cost[j] == dp_cost[best_j] and (j - P) < (best_j - P):
            best_j = j
    if best_j == -1:
        for j in range(max_cap, -1, -1):
            if dp_cost[j] != INF:
                best_j = j
                break
    if best_j == -1:
        return INF, []
    selected = []
    j = best_j
    for i in range(len(vehicles) - 1, -1, -1):
        v    = vehicles[i]
        cap  = v['capacidad']
        cost = v['costo']
        if j >= cap:
            prev_j = j - cap
            if (snapshots[i][prev_j] != INF and
                    snapshots[i][prev_j] + cost == snapshots[i + 1][j]):
                selected.append(v)
                j = prev_j
    trips     = []
    remaining = P
    for v in sorted(selected, key=lambda x: -x['capacidad']):
        load = min(v['capacidad'], remaining)
        trips.append({
            'conductor':          v['conductor'],
            'tipo':               v['tipo'],
            'capacidad':          v['capacidad'],
            'costo':              v['costo'],
            'costo_base':         v['costo'],
            'pallets_negociados': v.get('pallets_negociados', v['capacidad']),
            'costo_extra_pallet': v.get('costo_extra_pallet', 0),
            'vehicle_id':         v.get('vehicle_id', ''),
            'pallets_cargados':   max(1, load),
            'farms':              {},
        })
        remaining -= load
    return dp_cost[best_j], trips


def recalculate_variable_costs(trips):
    """Actualiza el costo de viajes con precio variable (ej. Yuber) segun pallets reales.
    También sincroniza pallets_cargados al load real (suma de farms), ya que el slot
    inicial del DP puede ser mayor que la carga efectiva tras restricciones de cuarteo."""
    for t in trips:
        actual = sum(fd['pallets'] for fd in t.get('farms', {}).values())
        if actual > 0 or t.get('farms') is not None:
            t['pallets_cargados'] = actual  # sincronizar al load real
        if t.get('costo_extra_pallet', 0) > 0:
            extra    = max(0, actual - t['pallets_negociados'])
            t['costo'] = t['costo_base'] + extra * t['costo_extra_pallet']
    return trips


# ── Etiqueta de pallets por talla ─────────────────────────────
def pallet_size_label(assigned_int_pallets, farm_total_pallets, pallets_by_size):
    if not pallets_by_size or farm_total_pallets <= 0:
        return str(assigned_int_pallets) + "P"
    ratio      = assigned_int_pallets / farm_total_pallets
    raw        = {ps: cnt * ratio for ps, cnt in pallets_by_size.items()}
    floors     = {ps: int(v) for ps, v in raw.items()}
    remainders = {ps: raw[ps] - floors[ps] for ps in raw}
    extra      = assigned_int_pallets - sum(floors.values())
    for ps, _ in sorted(remainders.items(), key=lambda x: -x[1])[:max(0, int(extra))]:
        floors[ps] += 1
    parts = [str(cnt) + "P (" + str(ps) + "c)"
             for ps, cnt in sorted(floors.items(), key=lambda x: -x[0]) if cnt > 0]
    return " + ".join(parts)


def format_pallets_by_size(pallets_by_size, total_pallets=0):
    """Texto tipo '30P x55c + 30P x44c' a partir de un dict {talla: cantidad}."""
    parts = [f"{cnt}P x{ps}c"
             for ps, cnt in sorted(pallets_by_size.items(), key=lambda x: -x[0]) if cnt > 0]
    if parts:
        return " + ".join(parts)
    return f"{int(total_pallets)}P" if total_pallets else '-'


# ── Distribución de fincas en viajes ─────────────────────────
def assign_farms_to_trips(farm_pallets, farm_cajas, trips):
    # Mínimo absoluto para fills en viajes con finca existente.
    # El objetivo es siempre llenar el camión más barato primero (optimizar peso).
    # Los fills pequeños son válidos: STA MARIA 4P, CHISPERO 4P, JUANA PIO 9P, etc.
    # Solo bloqueamos fills absurdos de 1-2P que no valen la parada en finca.
    FILL_MIN_PALLETS = 3

    sorted_farms = sorted(farm_pallets.items(), key=lambda x: -x[1])
    for trip in trips:
        trip['remaining'] = trip['pallets_cargados']
        trip['farms']     = {}
    for farm, fpallets in sorted_farms:
        rem_pallets = int(fpallets)
        rem_cajas   = farm_cajas[farm]
        for trip in trips:
            # Respetar restricciones conductor-finca (ej: Edwin no va a San Bartolo)
            conductor = trip.get('conductor', '')
            if farm in CONDUCTOR_FARM_RESTRICTIONS.get(conductor, set()):
                continue
            if trip['remaining'] > 0 and rem_pallets > 0:
                take = min(int(trip['remaining']), rem_pallets)
                min_cuarteo = CUARTEO_MIN_PALLETS.get(farm, CUARTEO_MIN_DEFAULT)
                if trip['farms']:
                    # Viaje con finca(s) existente: fill libre.
                    # Prioridad = llenar el camión al máximo (optimizar peso).
                    # Regla: bloquear solo si take < mínimo Y hay más pallets esperando
                    #        de esta finca (no son los últimos). Si son los últimos
                    #        pallets de la finca, siempre permitir (evita 2P huérfanos).
                    is_last_batch = (take >= rem_pallets)
                    if take < FILL_MIN_PALLETS and not is_last_batch:
                        continue
                    # Cuarteo: finca nueva en viaje con otra finca debe
                    # cumplir el mínimo siempre (sin excepción por is_last).
                    if farm not in trip['farms'] and take < min_cuarteo:
                        continue
                    # Chequeo fincas existentes: si alguna finca ya en el viaje
                    # está por debajo de su mínimo de cuarteo, no añadir más
                    # fincas a ese viaje (evita mezclas inválidas).
                    if farm not in trip['farms']:
                        _existing_ok_a = all(
                            trip['farms'][_ef]['pallets'] >= CUARTEO_MIN_PALLETS.get(_ef, CUARTEO_MIN_DEFAULT)
                            for _ef in trip['farms']
                        )
                        if not _existing_ok_a:
                            continue
                    # Evitar overage innecesario en camiones con tarifa variable.
                    # Si el fill supera los pallets negociados y no es el último
                    # lote de esta finca, limitar al espacio en tarifa plana.
                    # Ej: CHISPERO 20P + SALV 4P fill = 24P ($730k) es peor que
                    #     CHISPERO 20P ($630k) + SALV 11P viaje propio ($630k).
                    if trip.get('costo_extra_pallet', 0) > 0 and not is_last_batch:
                        _farms_loaded = sum(fd['pallets'] for fd in trip['farms'].values())
                        _pals_neg     = trip.get('pallets_negociados', trip.get('capacidad', 24))
                        if _farms_loaded + take > _pals_neg:
                            take = max(0, _pals_neg - _farms_loaded)
                            if take < FILL_MIN_PALLETS:
                                continue
                            # Re-chequear cuarteo con take ajustado: el overage puede
                            # haber reducido take de >min (pasó el chequeo) a <min.
                            if farm not in trip['farms'] and take < min_cuarteo:
                                continue
                else:
                    # Viaje vacío (primera finca): sí respetar mínimo de cuarteo.
                    # Si el slot DP es demasiado pequeño Y la finca tiene más pallets
                    # disponibles → saltarse este slot para que el mop-up la ubique
                    # en un camión adecuado. Previene JUANA PIO 1P en slot pequeño.
                    if take < min_cuarteo and rem_pallets > take:
                        continue
                rem_pallets -= take
                if rem_pallets == 0:
                    take_cajas = rem_cajas
                else:
                    ratio      = take / fpallets if fpallets > 0 else 0
                    take_cajas = min(int(round(ratio * farm_cajas[farm])), rem_cajas)
                trip['farms'][farm] = {'pallets': take, 'cajas': take_cajas}
                trip['remaining']  -= take
                rem_cajas          -= take_cajas

    # ── Post-proceso: eliminar fills absurdos de 1-2P ──────────────────────
    # Solo se eliminan fills menores a FILL_MIN_PALLETS (1-2P que no valen la parada).
    # Fills válidos como STA MARIA 4P, JUANA PIO 9P se mantienen.
    for trip in trips:
        if len(trip.get('farms', {})) < 2:
            continue
        violating = [
            f for f, fd in list(trip['farms'].items())
            if fd['pallets'] < 2   # Solo eliminar fills patológicos de 1P
        ]
        for f in violating:
            trip['farms'].pop(f)
    return trips


# ── Relleno combinado Chigorodó→Apartadó ─────────────────────
def _combined_fill(chigorodo_trips, apart_route_data):
    """
    Rellena la capacidad sobrante de camiones Chigorodó con fincas de Apartadó.
    Yuber: cobra $100.000 extra por la entrada a Apartadó.
    Demetrio/Edwin: tarifa plana, sin costo adicional por mezclar zonas.
    El recorrido es siempre en ruta: Chigorodó primero, luego Apartadó.
    Modifica apart_route_data in-place reduciendo los pallets consumidos.
    """
    ENTRADA_APARTADO  = 100_000
    CONDUCTORS_TARDE  = {'YUBER', 'DEMETRIO', 'EDWIN'}

    for trip in chigorodo_trips:
        conductor = trip.get('conductor', '')
        if conductor not in CONDUCTORS_TARDE:
            continue
        spare = trip['capacidad'] - trip['pallets_cargados']
        if spare <= 0:
            continue

        added_any = False
        for farm in sorted(apart_route_data, key=lambda f: -apart_route_data[f]['pallets']):
            if spare <= 0:
                break
            fdata = apart_route_data[farm]
            if fdata['pallets'] <= 0:
                continue
            # Restricción global: conductor no puede ir a ciertas fincas
            if farm in CONDUCTOR_FARM_RESTRICTIONS.get(conductor, set()):
                continue
            # Fill Chigorodó→Apartadó: mismo criterio de fill libre.
            # Objetivo: llenar el camión al máximo con lo que haya disponible.
            # Solo rechazar fills absurdos de 1-2P.
            _cuarteo_min_cf = CUARTEO_MIN_PALLETS.get(farm, CUARTEO_MIN_DEFAULT)
            take_p = min(spare, fdata['pallets'])
            # Si el viaje ya lleva otra finca, respetar cuarteo mínimo.
            _new_farm_cf = farm not in trip.get('farms', {})
            if trip.get('farms') and _new_farm_cf and take_p < _cuarteo_min_cf:
                continue
            # Chequeo fincas existentes: si alguna finca ya cargada está bajo
            # su mínimo, no mezclar con otra finca nueva (evita cuarteo inválido).
            if _new_farm_cf and trip.get('farms'):
                _existing_ok_cf = all(
                    trip['farms'][_ef]['pallets'] >= CUARTEO_MIN_PALLETS.get(_ef, CUARTEO_MIN_DEFAULT)
                    for _ef in trip['farms']
                )
                if not _existing_ok_cf:
                    continue
            if take_p < 3:  # bloquear fills patológicos de 1-2P
                continue
            # Cajas proporcionales (exactas si es el ultimo trozo)
            if take_p == fdata['pallets']:
                take_c = fdata['cajas']
            else:
                take_c = min(int(round(take_p / fdata['pallets'] * fdata['cajas'])), fdata['cajas'])

            existing = trip['farms'].get(farm, {'pallets': 0, 'cajas': 0})
            trip['farms'][farm] = {
                'pallets': existing['pallets'] + take_p,
                'cajas':   existing['cajas']   + take_c,
            }
            trip['pallets_cargados'] += take_p
            fdata['pallets'] -= take_p
            fdata['cajas']   -= take_c
            spare             -= take_p
            added_any          = True

        if added_any and conductor == 'YUBER':
            # Yuber: recalcular costo con entrada Apartadó + extra pallets
            extra      = max(0, trip['pallets_cargados'] - trip['pallets_negociados'])
            trip['costo'] = (trip['costo_base']
                             + ENTRADA_APARTADO
                             + extra * trip['costo_extra_pallet'])
        # Demetrio/Edwin: tarifa plana, costo no cambia


# ── Etiquetas Mediodía / Tarde ────────────────────────────────
def label_trip_times(trips):
    """
    Asigna 'hora' = 'Mediodía' o 'Tarde' a cada viaje de exportación.
    - Yuber: siempre 'Tarde'
    - Demetrio/Edwin con 2+ viajes en el día:
        El viaje con una sola zona (zona pura) y más pallets = 'Mediodía'.
        El resto = 'Tarde' (incluyendo los que mezclan Chigorodó→Apartadó).
    - Demetrio/Edwin con 1 solo viaje → 'Tarde'
    """
    export_trips = [t for t in trips if t.get('trip_type') == 'export']

    by_conductor = {}
    for t in export_trips:
        by_conductor.setdefault(t.get('conductor', ''), []).append(t)

    for conductor, ctrips in by_conductor.items():
        if conductor not in ('DEMETRIO', 'EDWIN'):
            for t in ctrips:
                t['hora'] = 'Tarde'
            continue

        if len(ctrips) >= 2:
            # Ordena: zona pura (1 zona) primero, luego pallets desc.
            # El trip más lleno y sin mezcla de zonas = Mediodía.
            # Trips con mezcla Chigorodó→Apartadó siempre quedan como Tarde.
            def _key(t):
                n_zones = len({FARM_ZONES.get(f, t.get('zone', ''))
                               for f in t.get('farms', {})})
                return (n_zones, -t.get('pallets_cargados', 0))
            ctrips_sorted = sorted(ctrips, key=_key)
            ctrips_sorted[0]['hora'] = 'Mediodía'
            for t in ctrips_sorted[1:]:
                t['hora'] = 'Tarde'
        else:
            ctrips[0]['hora'] = 'Tarde'

    return trips


# ── Optimizacion diaria ───────────────────────────────────────
def _optimize_phase(phase_orders, unavailable_vehicle_ids=None, enable_combined_fill=True, min_pallets=5, zone_order=None):
    if unavailable_vehicle_ids is None:
        unavailable_vehicle_ids = set()

    route_groups = {}
    for farm, port_data in phase_orders.items():
        zone = FARM_ZONES.get(farm, 'APARTADO')
        for port, data in port_data.items():
            key = (zone, port)
            if key not in route_groups:
                route_groups[key] = {}
            route_groups[key][farm] = {
                'pallets':         data['pallets'],
                'cajas':           data['cajas'],
                'pallets_by_size': data.get('pallets_by_size', {}),
            }

    all_trips        = []
    used_vehicle_ids = set()

    _route_avail = {
        rk: [v for v in VEHICLES_BY_ROUTE.get(rk, [])
             if v.get('vehicle_id', '') not in unavailable_vehicle_ids]
        for rk in route_groups
    }

    def _shared_demand(route_key):
        my_vids    = {v['vehicle_id'] for v in _route_avail[route_key]}
        other_vids = set()
        for k in route_groups:
            if k != route_key:
                other_vids |= {v['vehicle_id'] for v in _route_avail[k]}
        excl_cap = sum(v['capacidad'] for v in _route_avail[route_key]
                       if v['vehicle_id'] not in other_vids)
        total_p  = sum(d['pallets'] for d in route_groups[route_key].values())
        # Rutas sin capacidad exclusiva dependen 100% de vehículos compartidos
        # → deben procesarse primero para garantizar que no pierdan cajas.
        if excl_cap == 0:
            return float('inf')
        return max(0, total_p - excl_cap)

    # Desempate de zonas.  Modo normal: CHIGORODO primero para que el
    # combined-fill (Chigorodó → Apartadó) funcione correctamente.
    # Modo APARTADO_FIRST: APARTADO va primero para que EDWIN quede libre
    # para DF/Chispero en lugar de irse a JUANA PIO.
    if zone_order == 'APARTADO_FIRST':
        ZONE_PRIORITY = {'APARTADO': 1, 'CHIGORODO': 0}
    else:
        ZONE_PRIORITY = {'CHIGORODO': 1, 'APARTADO': 0}
    sorted_routes = sorted(
        route_groups.items(),
        key=lambda x: (_shared_demand(x[0]), ZONE_PRIORITY.get(x[0][0], 0)),
        reverse=True,
    )

    combined_done   = set()  # puertos donde ya se intento relleno combinado
    processed_routes = set()  # rutas (zona, puerto) ya procesadas

    for (zone, port), farm_data in sorted_routes:
        # Filtrar fincas con pallets reales (pueden haberse reducido por relleno combinado)
        farm_pallets = {f: d['pallets'] for f, d in farm_data.items() if d['pallets'] > 0}
        farm_cajas   = {f: d['cajas']   for f, d in farm_data.items() if d['pallets'] > 0}
        farm_psize   = {f: d.get('pallets_by_size', {}) for f, d in farm_data.items() if d['pallets'] > 0}
        total = sum(farm_pallets.values())
        if total <= 0:
            continue

        def avail(v):
            vid = v.get('vehicle_id', '')
            return vid not in unavailable_vehicle_ids and vid not in used_vehicle_ids

        vehicles = [v for v in VEHICLES_BY_ROUTE.get((zone, port), []) if avail(v)]
        if not vehicles:
            continue

        def tag(trips_list, dest, ttype):
            for t in trips_list:
                t['zone']               = zone
                t['destination']        = dest
                t['trip_type']          = ttype
                t['farm_pallets_size']  = farm_psize
                t['farm_total_pallets'] = farm_pallets

        # ── Separar fincas restringidas de libres ─────────────────
        # Si hay fincas que ciertos conductores no pueden visitar (ej: Edwin → SB),
        # optimizar primero esas fincas con vehículos elegibles, luego las demás.
        restr_farms   = {f for f in farm_pallets
                         if any(f in CONDUCTOR_FARM_RESTRICTIONS.get(v['conductor'], set())
                                for v in vehicles)}
        eligible_veh  = [v for v in vehicles
                         if not any(f in CONDUCTOR_FARM_RESTRICTIONS.get(v['conductor'], set())
                                    for f in restr_farms)]
        ineligible_veh = [v for v in vehicles if v not in eligible_veh]

        if restr_farms and (ineligible_veh or len(farm_pallets) > len(restr_farms)):
            # Sub-optimización A: fincas restringidas → solo vehículos elegibles
            # Nota: también se activa cuando todas las fincas son restringidas pero
            # hay fincas libres adicionales — así SAN BARTOLO y JUANA PIO siempre
            # obtienen slots separados y no se mezclan en un slot 24P+4P inutilizable.
            rp = {f: farm_pallets[f] for f in restr_farms}
            rc = {f: farm_cajas[f]   for f in restr_farms}
            total_r = sum(rp.values())
            cost_r, trips_r = min_cost_assignment_bounded(total_r, eligible_veh)
            trips_r = assign_farms_to_trips(rp, rc, trips_r)
            recalculate_variable_costs(trips_r)
            tag(trips_r, port, 'export')
            used_r = {t.get('vehicle_id', '') for t in trips_r}

            # Relleno de capacidad sobrante en trips_r con fincas libres.
            # Ejemplo: DEMETRIO lleva STA MARIA 16P (cap=24P) → 8P libres que
            # pueden recibir DOÑA FRANCIA antes de que EDWIN tome el resto.
            fp2 = {f: farm_pallets[f] for f in farm_pallets if f not in restr_farms}
            fc2 = {f: farm_cajas[f]   for f in farm_pallets if f not in restr_farms}
            fp2_fill = dict(fp2)  # copia mutable para el relleno
            fc2_fill = dict(fc2)
            for _farm in sorted(fp2_fill, key=lambda f: -fp2_fill[f]):
                _rem_p  = fp2_fill[_farm]
                _rem_c  = fc2_fill[_farm]
                _orig_p = fp2[_farm]
                _cuarteo_min = CUARTEO_MIN_PALLETS.get(_farm, CUARTEO_MIN_DEFAULT)
                for _t in trips_r:
                    _cond = _t.get('conductor', '')
                    if _farm in CONDUCTOR_FARM_RESTRICTIONS.get(_cond, set()):
                        continue
                    _spare = _t.get('capacidad', 0) - _t.get('pallets_cargados', 0)
                    if _spare <= 0 or _rem_p <= 0:
                        continue
                    _take = min(_spare, _rem_p)
                    _is_last = (_take >= _rem_p)
                    # Respetar cuarteo mínimo: si la finca no está ya en el viaje
                    # y el take es menor que el mínimo requerido, solo permitir si
                    # es el último batch (todos los pallets restantes de la finca).
                    _already = _farm in _t.get('farms', {})
                    # Cuarteo estricto: finca nueva en viaje con otra finca
                    # siempre debe cumplir el mínimo, sin importar si son
                    # los últimos pallets disponibles de esa finca.
                    if not _already and _take < _cuarteo_min:
                        continue
                    # Chequeo fincas existentes: si alguna ya cargada está bajo
                    # su mínimo de cuarteo, no añadir más fincas a ese viaje.
                    if not _already and _t.get('farms', {}):
                        _existing_ok_iz = all(
                            _t['farms'][_ef]['pallets'] >= CUARTEO_MIN_PALLETS.get(_ef, CUARTEO_MIN_DEFAULT)
                            for _ef in _t['farms']
                        )
                        if not _existing_ok_iz:
                            continue
                    if _take < 2:  # Evitar fills de 1P que luego se eliminarían
                        continue
                    _ratio = _take / _orig_p if _orig_p > 0 else 0
                    _take_c = min(int(round(_ratio * fc2[_farm])), _rem_c)
                    if not _already:
                        _t['farms'][_farm] = {'pallets': 0, 'cajas': 0}
                    _t['farms'][_farm]['pallets'] += _take
                    _t['farms'][_farm]['cajas']   += _take_c
                    _t['pallets_cargados'] = _t.get('pallets_cargados', 0) + _take
                    _rem_p -= _take
                    _rem_c -= _take_c
                fp2_fill[_farm] = _rem_p
                fc2_fill[_farm] = _rem_c
            recalculate_variable_costs(trips_r)

            # Sub-optimización B: fincas libres RESTANTES → vehículos restantes (incl. Edwin)
            fp2 = {f: p for f, p in fp2_fill.items() if p > 0}
            fc2 = {f: fc2_fill[f] for f in fp2}
            veh2 = [v for v in vehicles if v.get('vehicle_id', '') not in used_r]
            total_f = sum(fp2.values())
            if total_f > 0 and veh2:
                cost_f, trips_f = min_cost_assignment_bounded(total_f, veh2)
                trips_f = assign_farms_to_trips(fp2, fc2, trips_f)
                recalculate_variable_costs(trips_f)
                tag(trips_f, port, 'export')
            else:
                trips_f = []

            trips_a  = trips_r + trips_f
            cost_a   = sum(t['costo'] for t in trips_r) + sum(t['costo'] for t in trips_f)
        else:
            cost_a, trips_a = min_cost_assignment_bounded(total, vehicles)
            trips_a = assign_farms_to_trips(farm_pallets, farm_cajas, trips_a)
            recalculate_variable_costs(trips_a)
            tag(trips_a, port, 'export')

        # ── Mop-up: pallets que quedaron sin asignar por restricción de cuarteo ──
        # Calcular residuos reales por finca (pedido - asignado en trips_a)
        assigned_by_farm = {}
        for _t in trips_a:
            for _f, _fd in _t.get('farms', {}).items():
                assigned_by_farm[_f] = assigned_by_farm.get(_f, 0) + _fd['pallets']
        for _farm, _total_p in farm_pallets.items():
            _residual_p = _total_p - assigned_by_farm.get(_farm, 0)
            if _residual_p <= 0:
                continue
            if _residual_p < min_pallets:
                # Residual pequeño: intentar añadirlo a un viaje que ya lleva esta
                # finca y tiene capacidad física sobrante (no arranca viaje nuevo).
                # Calculamos cajas proporcionales restantes para esta finca.
                _cajas_already = sum(
                    _t.get('farms', {}).get(_farm, {}).get('cajas', 0) for _t in trips_a)
                _res_cajas_sm = max(0, int(farm_cajas.get(_farm, 0)) - _cajas_already)
                for _t in trips_a:
                    if _farm not in _t.get('farms', {}):
                        continue
                    if _farm in CONDUCTOR_FARM_RESTRICTIONS.get(_t.get('conductor', ''), set()):
                        continue
                    _phys_spare = _t.get('capacidad', 0) - _t.get('pallets_cargados', 0)
                    if _phys_spare <= 0:
                        continue
                    _add = min(_phys_spare, _residual_p)
                    _t['farms'][_farm]['pallets']     += _add
                    _t['farms'][_farm]['cajas']       += _res_cajas_sm  # todos los cajas restantes
                    _t['pallets_cargados']             = _t.get('pallets_cargados', 0) + _add
                    _residual_p                       -= _add
                    break
                continue  # si quedó residual aún, lo maneja inter-day
            # Verificar restricción conductor de Yuber (en general no tiene)
            if _farm in CONDUCTOR_FARM_RESTRICTIONS.get('YUBER', set()):
                continue
            # Buscar vehículos disponibles para este residuo (preferir Yuber)
            # Excluir también los vehículos ya comprometidos en trips_a (mismo turno)
            _vids_in_trips = {t.get('vehicle_id', '') for t in trips_a}
            _mop_veh = [v for v in vehicles
                        if v.get('vehicle_id', '') not in unavailable_vehicle_ids
                        and v.get('vehicle_id', '') not in _vids_in_trips
                        and not (_farm in CONDUCTOR_FARM_RESTRICTIONS.get(v.get('conductor',''), set()))]
            _res_cajas = farm_cajas.get(_farm, 0) - sum(
                _t.get('farms', {}).get(_farm, {}).get('cajas', 0) for _t in trips_a)
            _res_cajas = max(0, int(_res_cajas))
            if not _mop_veh:
                # No hay vehículos libres: reutilizar viajes vacíos (slots del DP
                # que quedaron sin fincas porque todas las fincas saltaron el slot
                # pequeño gracias a la protección de cuarteo mínimo).
                # Ejemplo: YUBER_1 con slot 4P y 0 pallets asignados puede recibir
                # JUANA PIO 13P si expandimos su slot a la capacidad real (24P).
                for _t in trips_a:
                    if _t.get('farms', {}):
                        continue  # este viaje ya tiene fincas — no reutilizar
                    if _farm in CONDUCTOR_FARM_RESTRICTIONS.get(_t.get('conductor',''), set()):
                        continue
                    # Expandir slot al mínimo necesario y asignar la finca
                    _slot = min(_t['capacidad'], _residual_p)
                    _t['pallets_cargados'] = _slot
                    _t['remaining']        = _slot
                    _t['farms']            = {}
                    _assign_r = assign_farms_to_trips(
                        {_farm: _residual_p}, {_farm: _res_cajas}, [_t])
                    recalculate_variable_costs(_assign_r)
                    tag(_assign_r, port, 'export')
                    # El viaje ya está en trips_a; solo actualizar costo
                    cost_a += _t['costo']
                    _residual_p = 0
                    break
                continue  # si aún quedan pallets, inter-day los maneja
            _, _mop_trips = min_cost_assignment_bounded(_residual_p, _mop_veh)
            _mop_trips = assign_farms_to_trips({_farm: _residual_p}, {_farm: _res_cajas}, _mop_trips)
            recalculate_variable_costs(_mop_trips)
            tag(_mop_trips, port, 'export')
            trips_a.extend(_mop_trips)
            cost_a += sum(_t['costo'] for _t in _mop_trips)

        consol_candidates = [f for f in farm_pallets if f in CONSOLIDACION_ROUTES]
        if not consol_candidates:
            for t in trips_a:
                used_vehicle_ids.add(t.get('vehicle_id', ''))
            all_trips.extend(trips_a)
        else:
            consol_trips = []
            consol_cost  = 0
            consol_used  = set()
            for farm in consol_candidates:
                cv = [v for v in CONSOLIDACION_ROUTES[farm]
                      if v.get('vehicle_id') not in unavailable_vehicle_ids
                      and v.get('vehicle_id') not in used_vehicle_ids
                      and v.get('vehicle_id') not in consol_used]
                if not cv:
                    continue
                fp = {farm: farm_pallets[farm]}
                fc = {farm: farm_cajas[farm]}
                cc, ct = min_cost_assignment_bounded(farm_pallets[farm], cv)
                consol_cost += cc
                ct = assign_farms_to_trips(fp, fc, ct)
                tag(ct, 'DONA FRANCIA', 'consolidacion')
                for t in ct:
                    consol_used.add(t.get('vehicle_id', ''))
                consol_trips.extend(ct)
            vehicles_b     = [v for v in vehicles if v.get('vehicle_id') not in consol_used]
            cost_b_exp, trips_b = min_cost_assignment_bounded(total, vehicles_b)
            trips_b = assign_farms_to_trips(farm_pallets, farm_cajas, trips_b)
            recalculate_variable_costs(trips_b)
            tag(trips_b, port, 'export')
            cost_b       = consol_cost + cost_b_exp
            chosen_trips = consol_trips + trips_b if cost_b < cost_a else trips_a
            for t in chosen_trips:
                used_vehicle_ids.add(t.get('vehicle_id', ''))
            all_trips.extend(chosen_trips)

        processed_routes.add((zone, port))

        # Relleno combinado: si acabamos de procesar Chigorodo y Apartado aun
        # no ha sido procesado, aprovechar capacidad sobrante de Yuber/Edwin
        # para recoger fincas de Apartado (reduce lo que Apartado tendra que cubrir).
        # Si Apartado ya fue procesado primero, NO hacer relleno para evitar
        # contar cajas dos veces.
        if enable_combined_fill and zone == 'CHIGORODO' and port not in combined_done:
            apart_key = ('APARTADO', port)
            if apart_key in route_groups and apart_key not in processed_routes:
                combined_done.add(port)
                chigorodo_trips = [t for t in all_trips
                                   if t.get('zone') == 'CHIGORODO'
                                   and t.get('destination') == port
                                   and t.get('trip_type') == 'export']
                _combined_fill(chigorodo_trips, route_groups[apart_key])

    return all_trips


# ── Sugerencias de consolidacion entre dias ──────────────────

# ── Helpers para split mediodía / tarde ──────────────────────
def _cap_to_mediodia(day_orders):
    """Demanda capada a FARM_MEDIODIA_MAX: solo lo que está listo al mediodía."""
    result = {}
    for farm, port_data in day_orders.items():
        cap = FARM_MEDIODIA_MAX.get(farm, float('inf'))
        if cap <= 0:
            continue
        for port, data in port_data.items():
            if data.get('pallets', 0) <= 0:
                continue
            m_p = min(data['pallets'], cap)
            if m_p <= 0:
                continue
            ratio = m_p / data['pallets']
            result.setdefault(farm, {})[port] = {
                'pallets':         m_p,
                'cajas':           int(round(ratio * data['cajas'])),
                'pallets_by_size': {k: max(1, int(round(v * ratio)))
                                    for k, v in data.get('pallets_by_size', {}).items()
                                    if v > 0},
            }
    return result


def _compute_tarde_demand(day_orders, mediodia_trips):
    """Demanda restante = pedido original menos lo ya asignado en la fase mediodía."""
    assigned = {}
    for t in mediodia_trips:
        if t.get('trip_type') != 'export':
            continue
        port = t.get('destination', '')
        for farm, fdata in t['farms'].items():
            key = (farm, port)
            assigned.setdefault(key, {'pallets': 0, 'cajas': 0})
            assigned[key]['pallets'] += fdata.get('pallets', 0)
            assigned[key]['cajas']   += fdata.get('cajas', 0)

    result = {}
    for farm, port_data in day_orders.items():
        for port, data in port_data.items():
            asgn  = assigned.get((farm, port), {'pallets': 0, 'cajas': 0})
            rem_p = data['pallets'] - asgn['pallets']
            rem_c = data['cajas']   - asgn['cajas']
            if rem_p <= 0:
                continue
            ratio = rem_p / data['pallets'] if data['pallets'] > 0 else 0
            result.setdefault(farm, {})[port] = {
                'pallets':         rem_p,
                'cajas':           max(0, rem_c),
                'pallets_by_size': {k: max(0, int(round(v * ratio)))
                                    for k, v in data.get('pallets_by_size', {}).items()},
            }
    return result


# ── Optimizacion diaria (tres opciones: elige la más barata) ────────
def optimize_day(day_orders, unavailable_vehicle_ids=None, relaxed=False, cap_mediodia=False):
    """
    Evalúa 3 estrategias de despacho y elige la de menor costo:
      A) Mediodía con caps operativos (FARM_MEDIODIA_MAX) + tarde
      B) Mediodía libre sin caps (el DP decide cuánto va al mediodía) + tarde
      C) Sin mediodía — todo en tarde (Edwin/Demetrio solo 1 slot cada uno)
    El mediodía NO es obligatorio: solo se usa si genera un plan más barato.
    """
    if unavailable_vehicle_ids is None:
        unavailable_vehicle_ids = set()

    VIAJE2_VIDS = {'DEMETRIO_PATINETA_2', 'EDWIN_MULA_2'}
    VIAJE1_VIDS = {'DEMETRIO_PATINETA', 'EDWIN_MULA'}
    TARDE_MIN   = 0 if relaxed else 5

    def _mediodia_min(cap):
        if relaxed:
            return 0
        return int(cap * (0.60 if cap <= 18 else 0.50))

    def _run_two_phase(mediodia_demand, zone_order=None):
        """Fase mediodía con `mediodia_demand` + tarde con el resto."""
        md = _optimize_phase(
            mediodia_demand,
            unavailable_vehicle_ids=unavailable_vehicle_ids | VIAJE2_VIDS,
            enable_combined_fill=False,
            zone_order=zone_order,
        )
        # Descartar mediodías con menos del 50%/60% de carga
        md = [t for t in md if t.get('trip_type') != 'export'
              or (t.get('pallets_cargados', 0) > 0
                  and t.get('pallets_cargados', 0) >= _mediodia_min(t.get('capacidad', 24)))]
        for t in md:
            t['hora'] = 'Mediodía'

        md_vids    = {t['vehicle_id'] for t in md}
        unused_v1  = VIAJE1_VIDS - md_vids - unavailable_vehicle_ids
        tarde_dem  = _compute_tarde_demand(day_orders, md)
        td = _optimize_phase(
            tarde_dem,
            unavailable_vehicle_ids=unavailable_vehicle_ids | md_vids | unused_v1,
            enable_combined_fill=True,
            min_pallets=TARDE_MIN,
            zone_order=zone_order,
        )
        td = [t for t in td if t.get('trip_type') != 'export'
              or (t.get('pallets_cargados', 0) > 0
                  and t.get('pallets_cargados', 0) >= TARDE_MIN)]
        for t in td:
            t['hora'] = 'Tarde'
        return md + td

    def _run_tarde_only():
        """Sin mediodía: VIAJE1 bloqueado, cada conductor solo 1 slot tarde."""
        td = _optimize_phase(
            day_orders,
            unavailable_vehicle_ids=unavailable_vehicle_ids | VIAJE1_VIDS,
            enable_combined_fill=True,
            min_pallets=TARDE_MIN,
        )
        td = [t for t in td if t.get('trip_type') != 'export'
              or (t.get('pallets_cargados', 0) > 0
                  and t.get('pallets_cargados', 0) >= TARDE_MIN)]
        for t in td:
            t['hora'] = 'Tarde'
        return td

    def _export_cost(trips):
        return sum(t.get('costo', 0) for t in trips if t.get('trip_type') == 'export')

    def _export_pallets(trips):
        return sum(t.get('pallets_cargados', 0) for t in trips if t.get('trip_type') == 'export')

    # Opción A: mediodía con caps operativos + tarde  (CHIGORODO primero)
    opt_a = _run_two_phase(_cap_to_mediodia(day_orders))

    # Opción C: sin mediodía, todo tarde
    opt_c = _run_tarde_only()

    # Opción D: APARTADO primero + caps mediodía
    # EDWIN → DF/Chispero (APARTADO $600k flat), JUANA PIO → YUBER
    # DEMETRIO → SALVAMENTO/lotes pequeños APARTADO ($550k) antes que SAN BARTOLO
    # Ganancia neta cuando ahorro APARTADO > extra en JUANA PIO.
    opt_d = _run_two_phase(_cap_to_mediodia(day_orders), zone_order='APARTADO_FIRST')

    if cap_mediodia:
        # Excluir B y E (mediodía sin caps): físicamente imposible superar FARM_MEDIODIA_MAX
        options = [opt_a, opt_c, opt_d]
    else:
        # Opción B: mediodía sin caps (flexible) + tarde  (CHIGORODO primero)
        opt_b = _run_two_phase(day_orders)
        # Opción E: APARTADO primero, mediodía sin caps
        opt_e = _run_two_phase(day_orders, zone_order='APARTADO_FIRST')
        options = [opt_a, opt_b, opt_c, opt_d, opt_e]

    # Elegir: tolerancia de 5P — si una opción es más barata y solo pierde ≤5P,
    # esos pallets van a inter-day al día siguiente (suele ser más económico).
    max_pals  = max(_export_pallets(o) for o in options)
    PALLET_TOLERANCE = 5
    best_opts = [o for o in options if _export_pallets(o) >= max_pals - PALLET_TOLERANCE]
    return min(best_opts, key=_export_cost)


def _day_metrics(day_orders, unavailable_vehicle_ids=None):
    """Devuelve (costo_total, cajas_perdidas) para un pedido de un dia."""
    trips = optimize_day(day_orders, unavailable_vehicle_ids=unavailable_vehicle_ids)
    export_trips = [t for t in trips if t.get('trip_type') == 'export']
    cost          = sum(t['costo'] for t in trips)
    cajas_shipped = sum(sum(f['cajas'] for f in t['farms'].values()) for t in export_trips)
    cajas_total   = sum(d['cajas'] for port_data in day_orders.values() for d in port_data.values())
    return cost, max(0, cajas_total - cajas_shipped)


def _split_blocks(entry):
    """Divide el pedido de una finca/puerto en bloques por talla de pallet."""
    pallets_by_size = entry.get('pallets_by_size', {})
    total_pallets   = entry['pallets']
    total_cajas     = entry['cajas']
    if not pallets_by_size or total_pallets <= 0:
        return [{'size': None, 'pallets': total_pallets, 'cajas': total_cajas}]
    items = sorted(pallets_by_size.items())
    blocks = []
    assigned_cajas = 0
    for idx, (size, count) in enumerate(items):
        if count <= 0:
            continue
        if idx == len(items) - 1:
            cajas = total_cajas - assigned_cajas
        else:
            cajas = int(round(total_cajas * (count / total_pallets)))
            assigned_cajas += cajas
        blocks.append({'size': size, 'pallets': count, 'cajas': cajas})
    return blocks


def _remove_blocks(day_orders, farm, port, blocks):
    """Copia day_orders sin los bloques indicados (restados de farm/port)."""
    new_orders = copy.deepcopy(day_orders)
    entry = new_orders.get(farm, {}).get(port)
    if not entry:
        return new_orders
    for b in blocks:
        entry['pallets'] -= b['pallets']
        entry['cajas']   -= b['cajas']
        if b['size'] is not None:
            psz = entry.get('pallets_by_size', {})
            psz[b['size']] = psz.get(b['size'], 0) - b['pallets']
            if psz[b['size']] <= 0:
                psz.pop(b['size'], None)
    if entry['pallets'] <= 0:
        del new_orders[farm][port]
        if not new_orders[farm]:
            del new_orders[farm]
    return new_orders


def _add_blocks(day_orders, farm, port, blocks):
    """Copia day_orders con los bloques indicados sumados a farm/port."""
    new_orders = copy.deepcopy(day_orders)
    if farm not in new_orders:
        new_orders[farm] = {}
    if port not in new_orders[farm]:
        new_orders[farm][port] = {'cajas': 0, 'pallets': 0, 'pallets_by_size': {}}
    entry = new_orders[farm][port]
    entry.setdefault('pallets_by_size', {})
    for b in blocks:
        entry['pallets'] += b['pallets']
        entry['cajas']   += b['cajas']
        if b['size'] is not None:
            entry['pallets_by_size'][b['size']] = entry['pallets_by_size'].get(b['size'], 0) + b['pallets']
    return new_orders


def suggest_consolidations(orders, unavailable_vehicle_ids_by_day=None, min_savings=30000):
    """
    Busca oportunidades de juntar (total o parcialmente) el pedido de una
    misma finca+puerto entre dos dias consecutivos para ahorrar costo de
    transporte, sin aumentar las cajas perdidas.
    Devuelve una lista de sugerencias ordenadas por ahorro descendente.
    """
    if unavailable_vehicle_ids_by_day is None:
        unavailable_vehicle_ids_by_day = {}

    sorted_days = sorted(orders.keys(), key=lambda d: DAY_ORDER.index(d) if d in DAY_ORDER else 99)
    suggestions = []

    for i in range(len(sorted_days) - 1):
        day_a, day_b = sorted_days[i], sorted_days[i + 1]
        unavail_a = unavailable_vehicle_ids_by_day.get(day_a, set())
        unavail_b = unavailable_vehicle_ids_by_day.get(day_b, set())
        orders_a, orders_b = orders[day_a], orders[day_b]

        cost_a0, loss_a0 = _day_metrics(orders_a, unavail_a)
        cost_b0, loss_b0 = _day_metrics(orders_b, unavail_b)
        base_cost = cost_a0 + cost_b0
        base_loss = loss_a0 + loss_b0

        farms_ports = set()
        for farm, pd_ in orders_a.items():
            for port in pd_:
                farms_ports.add((farm, port))
        for farm, pd_ in orders_b.items():
            for port in pd_:
                farms_ports.add((farm, port))

        for farm, port in farms_ports:
            for from_day, to_day, src_orders, dst_orders, unavail_src, unavail_dst in (
                (day_a, day_b, orders_a, orders_b, unavail_a, unavail_b),
                (day_b, day_a, orders_b, orders_a, unavail_b, unavail_a),
            ):
                entry = src_orders.get(farm, {}).get(port)
                if not entry or entry['pallets'] <= 0:
                    continue
                blocks = _split_blocks(entry)
                n = len(blocks)
                best = None
                for r in range(1, n + 1):
                    for combo in combinations(range(n), r):
                        move_blocks = [blocks[k] for k in combo]
                        new_src = _remove_blocks(src_orders, farm, port, move_blocks)
                        new_dst = _add_blocks(dst_orders, farm, port, move_blocks)
                        if from_day == day_a:
                            cost_a1, loss_a1 = _day_metrics(new_src, unavail_a)
                            cost_b1, loss_b1 = _day_metrics(new_dst, unavail_b)
                        else:
                            cost_b1, loss_b1 = _day_metrics(new_src, unavail_b)
                            cost_a1, loss_a1 = _day_metrics(new_dst, unavail_a)
                        new_cost = cost_a1 + cost_b1
                        new_loss = loss_a1 + loss_b1
                        if new_loss > base_loss:
                            continue
                        savings = base_cost - new_cost
                        if savings <= 0:
                            continue
                        if best is None or savings > best['savings']:
                            best = {'savings': savings, 'blocks': move_blocks, 'is_full': (r == n)}
                if best and best['savings'] >= min_savings:
                    suggestions.append({
                        'farm':     farm,
                        'port':     port,
                        'from_day': from_day,
                        'to_day':   to_day,
                        'blocks':   best['blocks'],
                        'pallets':  sum(b['pallets'] for b in best['blocks']),
                        'cajas':    sum(b['cajas'] for b in best['blocks']),
                        'is_full':  best['is_full'],
                        'savings':  best['savings'],
                    })

    suggestions.sort(key=lambda s: -s['savings'])
    return suggestions


def apply_consolidation(orders, suggestion):
    """Aplica una sugerencia de consolidacion, modificando 'orders' in-place."""
    farm, port   = suggestion['farm'], suggestion['port']
    from_day     = suggestion['from_day']
    to_day       = suggestion['to_day']
    blocks       = suggestion['blocks']
    orders[from_day] = _remove_blocks(orders[from_day], farm, port, blocks)
    orders[to_day]   = _add_blocks(orders[to_day], farm, port, blocks)
    return orders


# ── Consolidación entre días ──────────────────────────────────
def compute_inter_day_moves(orders):
    """
    Detecta y propone movimientos de pallets entre días consecutivos.
    Regla 1 – DIFERIMIENTO ADELANTE: pallets no enviados el día D van al día D+1
              si la finca tiene pedido en D+1.
    Regla 2 – DIFERIMIENTO ATRÁS: pallets no enviados en el último día posible
              (sin día siguiente con esa finca) se anticipan al día D-1.
    Regla 3 – ANTICIPACIÓN: pallets del día D+1 muy pocos (< 5P) se
              adelantan al día D si la finca tiene pedido en D.
    Retorna (adjusted_orders, moves_list).
    """
    ANTICIPATION_MAX = 5
    dias     = [d for d in DAY_ORDER if d in orders]
    adjusted = copy.deepcopy(orders)
    moves    = []

    # Paso 1: Diferimientos
    for i, dia in enumerate(dias):
        if i + 1 >= len(dias):
            break
        next_dia   = dias[i + 1]
        day_orders = orders.get(dia, {})
        if not day_orders:
            continue
        trips   = optimize_day(day_orders)
        shipped = {}
        for t in trips:
            if t.get('trip_type') != 'export':
                continue
            port = t.get('destination', '')
            for farm, fd in t['farms'].items():
                shipped[(farm, port)] = shipped.get((farm, port), 0) + fd['pallets']

        for farm, port_data in day_orders.items():
            if farm not in adjusted.get(next_dia, {}):
                continue
            for port, data in port_data.items():
                diff = data['pallets'] - shipped.get((farm, port), 0)
                if diff <= 0:
                    continue
                ratio       = diff / data['pallets'] if data['pallets'] > 0 else 0
                extra_cajas = max(1, int(round(ratio * data['cajas'])))
                adjusted[dia][farm][port]['pallets'] -= diff
                adjusted[dia][farm][port]['cajas']    = max(0,
                    adjusted[dia][farm][port]['cajas'] - extra_cajas)
                if adjusted[dia][farm][port]['pallets'] <= 0:
                    adjusted[dia][farm].pop(port, None)
                if not adjusted[dia].get(farm):
                    adjusted[dia].pop(farm, None)
                if port not in adjusted[next_dia].get(farm, {}):
                    adjusted[next_dia].setdefault(farm, {})[port] = {
                        'pallets': 0, 'cajas': 0, 'pallets_by_size': {}}
                adjusted[next_dia][farm][port]['pallets'] += diff
                adjusted[next_dia][farm][port]['cajas']   += extra_cajas
                moves.append({'type': 'diferimiento', 'farm': farm,
                              'from_day': dia, 'to_day': next_dia,
                              'pallets': diff, 'cajas': extra_cajas,
                              'reason': f'Sin enviar el {dia.title()} — pocas unidades sin viaje disponible'})

    # Paso 1b: Diferimiento ATRÁS — si el último día no puede despachar algo
    # y no hay día siguiente con esa finca, moverlo al día anterior
    last_dia = dias[-1]
    last_orders = orders.get(last_dia, {})
    if last_orders:
        trips_last = optimize_day(last_orders)
        shipped_last = {}
        for t in trips_last:
            if t.get('trip_type') != 'export': continue
            port = t.get('destination', '')
            for farm, fd in t['farms'].items():
                shipped_last[(farm, port)] = shipped_last.get((farm, port), 0) + fd['pallets']
        prev_dia = dias[-2] if len(dias) >= 2 else None
        if prev_dia:
            for farm, port_data in last_orders.items():
                if farm not in adjusted.get(prev_dia, {}):
                    continue
                for port, data in port_data.items():
                    diff = data['pallets'] - shipped_last.get((farm, port), 0)
                    if diff <= 0:
                        continue
                    ratio       = diff / data['pallets'] if data['pallets'] > 0 else 0
                    extra_cajas = max(1, int(round(ratio * data['cajas'])))
                    adjusted[last_dia][farm][port]['pallets'] -= diff
                    adjusted[last_dia][farm][port]['cajas']    = max(0,
                        adjusted[last_dia][farm][port]['cajas'] - extra_cajas)
                    if adjusted[last_dia][farm][port]['pallets'] <= 0:
                        adjusted[last_dia][farm].pop(port, None)
                    if not adjusted[last_dia].get(farm):
                        adjusted[last_dia].pop(farm, None)
                    if port not in adjusted[prev_dia].get(farm, {}):
                        adjusted[prev_dia].setdefault(farm, {})[port] = {
                            'pallets': 0, 'cajas': 0, 'pallets_by_size': {}}
                    adjusted[prev_dia][farm][port]['pallets'] += diff
                    adjusted[prev_dia][farm][port]['cajas']   += extra_cajas
                    moves.append({'type': 'diferimiento', 'farm': farm,
                                  'from_day': last_dia, 'to_day': prev_dia,
                                  'pallets': diff, 'cajas': extra_cajas,
                                  'reason': f'Sin enviar el {last_dia.title()} — se anticipan al {prev_dia.title()}'})

    # Paso 1c: Segunda pasada sobre pedidos YA ajustados para capturar residuos en cascada
    # Busca el PRIMER día futuro donde la finca tiene pedido (puede saltar días sin esa finca).
    # Ej: DOÑA FRANCIA 2P Martes → salta MIERCOLES (sin DOÑA FRANCIA) → llega a VIERNES.
    for i, dia in enumerate(dias):
        adj_day_ord = adjusted.get(dia, {})
        if not adj_day_ord:
            continue
        trips_adj   = optimize_day(adj_day_ord)
        shipped_adj = {}
        for t in trips_adj:
            if t.get('trip_type') != 'export':
                continue
            port = t.get('destination', '')
            for farm, fd in t['farms'].items():
                shipped_adj[(farm, port)] = shipped_adj.get((farm, port), 0) + fd['pallets']
        for farm, port_data in list(adj_day_ord.items()):
            # Buscar el próximo día (en cualquier posición futura) que tenga esta finca
            target_dia = None
            for future_dia in dias[i + 1:]:
                if farm in adjusted.get(future_dia, {}):
                    target_dia = future_dia
                    break
            if target_dia is None:
                continue  # No hay día futuro con esta finca — no se puede diferir
            for port, data in list(port_data.items()):
                diff = data['pallets'] - shipped_adj.get((farm, port), 0)
                if diff <= 0:
                    continue
                ratio       = diff / data['pallets'] if data['pallets'] > 0 else 0
                extra_cajas = max(1, int(round(ratio * data['cajas'])))
                adjusted[dia][farm][port]['pallets'] -= diff
                adjusted[dia][farm][port]['cajas']    = max(0,
                    adjusted[dia][farm][port]['cajas'] - extra_cajas)
                if adjusted[dia][farm][port]['pallets'] <= 0:
                    adjusted[dia][farm].pop(port, None)
                if not adjusted[dia].get(farm):
                    adjusted[dia].pop(farm, None)
                if port not in adjusted[target_dia].get(farm, {}):
                    adjusted[target_dia].setdefault(farm, {})[port] = {
                        'pallets': 0, 'cajas': 0, 'pallets_by_size': {}}
                adjusted[target_dia][farm][port]['pallets'] += diff
                adjusted[target_dia][farm][port]['cajas']   += extra_cajas
                moves.append({'type': 'diferimiento', 'farm': farm,
                              'from_day': dia, 'to_day': target_dia,
                              'pallets': diff, 'cajas': extra_cajas,
                              'reason': f'Residuo en cascada del {dia.title()} — se pasan al {target_dia.title()}'})

    # Paso 2: Anticipaciones
    for i, dia in enumerate(dias[:-1]):
        next_dia = dias[i + 1]
        for farm, port_data in list(adjusted.get(next_dia, {}).items()):
            for port, data in list(port_data.items()):
                p = data['pallets']
                if p <= 0 or p >= ANTICIPATION_MAX:
                    continue
                if farm not in adjusted.get(dia, {}):
                    continue
                c = data['cajas']
                if port not in adjusted[dia].get(farm, {}):
                    adjusted[dia].setdefault(farm, {})[port] = {
                        'pallets': 0, 'cajas': 0, 'pallets_by_size': {}}
                adjusted[dia][farm][port]['pallets'] += p
                adjusted[dia][farm][port]['cajas']   += c
                adjusted[next_dia][farm][port]['pallets'] -= p
                adjusted[next_dia][farm][port]['cajas']    = max(0,
                    adjusted[next_dia][farm][port]['cajas'] - c)
                if adjusted[next_dia][farm][port]['pallets'] <= 0:
                    adjusted[next_dia][farm].pop(port, None)
                if not adjusted[next_dia].get(farm):
                    adjusted[next_dia].pop(farm, None)
                moves.append({'type': 'anticipación', 'farm': farm,
                              'from_day': next_dia, 'to_day': dia,
                              'pallets': p, 'cajas': c,
                              'reason': f'Solo {p}P solos en {next_dia.title()} — se adelantan a {dia.title()}'})

    # ── Paso 3: Diferimiento deliberado (hill-climbing) ─────────────────
    # Intenta mover N pallets de la finca F del día D al día D+1
    # si reduce el costo total de ambos días. Restricciones:
    #   1. El día D debe seguir enviando TODOS sus pallets restantes.
    #   2. El día D+1 también debe poder enviar TODOS sus pallets.
    #   3. Máximo MAX_TRANSFER pallets por finca en cada par de días.
    # Esto evita sobrecargar un día y dejar pallets sin despachar.
    import copy as _copy2

    MAX_TRANSFER    = 12   # máximo pallets a mover por finca/día-par
    MAX_ITER_DD     = 2    # pasadas máximas

    def _day_result(day_orders_dict):
        """Retorna (pallets_enviados, costo_export) para un día."""
        if not day_orders_dict:
            return 0, 0
        ts = optimize_day(day_orders_dict)
        exp = [t for t in ts if t.get('trip_type') == 'export']
        return (sum(t['pallets_cargados'] for t in exp),
                sum(t['costo']           for t in exp))

    def _demand(day_orders_dict):
        return sum(d.get('pallets', 0)
                   for fd in day_orders_dict.values()
                   for d in fd.values())

    for _pass in range(MAX_ITER_DD):
        _improved = False
        for i in range(len(dias) - 1):
            dia      = dias[i]
            next_dia = dias[i + 1]
            adj_dia  = adjusted.get(dia, {})
            adj_nxt  = adjusted.get(next_dia, {})
            if not adj_dia or not adj_nxt:
                continue

            common_farms = [f for f in adj_dia if f in adj_nxt]
            if not common_farms:
                continue

            _ship_d, _cost_d = _day_result(adj_dia)
            _ship_n, _cost_n = _day_result(adj_nxt)
            # Si alguno ya no envía todo, no tocar este par
            if _ship_d < _demand(adj_dia) or _ship_n < _demand(adj_nxt):
                continue
            base_cost = _cost_d + _cost_n

            for farm in common_farms:
                for port, pdata in list(adjusted.get(dia, {}).get(farm, {}).items()):
                    total_p = pdata.get('pallets', 0)
                    total_c = pdata.get('cajas', 0)
                    if total_p <= 0:
                        continue

                    best_n    = 0
                    best_cost = base_cost
                    cap_n     = min(total_p, MAX_TRANSFER)

                    for n in range(1, cap_n + 1):
                        ratio   = n / total_p if total_p > 0 else 0
                        n_cajas = max(1, int(round(ratio * total_c)))

                        _adj_d = _copy2.deepcopy(adjusted.get(dia, {}))
                        _adj_n = _copy2.deepcopy(adjusted.get(next_dia, {}))

                        # Restar del día D
                        _adj_d[farm][port]['pallets'] -= n
                        _adj_d[farm][port]['cajas']    = max(0,
                            _adj_d[farm][port]['cajas'] - n_cajas)
                        if _adj_d[farm][port]['pallets'] <= 0:
                            _adj_d[farm].pop(port, None)
                        if not _adj_d.get(farm):
                            _adj_d.pop(farm, None)

                        # Sumar al día D+1
                        if port not in _adj_n.get(farm, {}):
                            _adj_n.setdefault(farm, {})[port] = {
                                'pallets': 0, 'cajas': 0, 'pallets_by_size': {}}
                        _adj_n[farm][port]['pallets'] += n
                        _adj_n[farm][port]['cajas']   += n_cajas

                        # Validar: ambos días envían TODA su demanda
                        _sh_d, _co_d = _day_result(_adj_d)
                        _sh_n, _co_n = _day_result(_adj_n)
                        dem_d = _demand(_adj_d)
                        dem_n = _demand(_adj_n)
                        if _sh_d < dem_d or _sh_n < dem_n:
                            continue  # no envía todo → rechazar

                        new_cost = _co_d + _co_n
                        if new_cost < best_cost:
                            best_cost = new_cost
                            best_n    = n

                    if best_n > 0:
                        ratio   = best_n / total_p if total_p > 0 else 0
                        n_cajas = max(1, int(round(ratio * total_c)))
                        adjusted[dia][farm][port]['pallets'] -= best_n
                        adjusted[dia][farm][port]['cajas']    = max(0,
                            adjusted[dia][farm][port]['cajas'] - n_cajas)
                        if adjusted[dia][farm][port]['pallets'] <= 0:
                            adjusted[dia][farm].pop(port, None)
                        if not adjusted[dia].get(farm):
                            adjusted[dia].pop(farm, None)
                        if port not in adjusted[next_dia].get(farm, {}):
                            adjusted[next_dia].setdefault(farm, {})[port] = {
                                'pallets': 0, 'cajas': 0, 'pallets_by_size': {}}
                        adjusted[next_dia][farm][port]['pallets'] += best_n
                        adjusted[next_dia][farm][port]['cajas']   += n_cajas
                        moves.append({
                            'type':      'diferimiento_deliberado',
                            'farm':      farm,
                            'farm_name': farm,
                            'from_day':  dia,
                            'to_day':    next_dia,
                            'pallets':   best_n,
                            'cajas':     n_cajas,
                            'ahorro':    base_cost - best_cost,
                            'reason':    (f'Diferimiento deliberado: {best_n}P de {farm} '
                                         f'{dia.title()}→{next_dia.title()} '
                                         f'ahorra ${base_cost - best_cost:,}'),
                        })
                        base_cost = best_cost
                        _improved = True
        if not _improved:
            break

    return adjusted, moves



# ── Plan original con carry-forward de residuos pequeños ─────────────────────
# Cuando una finca tiene menos de _MIN_DISPATCH_ORIG pallets solos en un viaje,
# no se despacha ese día — esos pallets se suman al día siguiente.
# Esto evita escenarios irreales como DEMETRIO con 1 pallet.
_MIN_DISPATCH_ORIG = 10  # mínimo realista: < 10P solo → difiere al día siguiente

def _merge_carry(base_orders, carry):
    """Añade pallets de carry-forward al pedido del siguiente día."""
    merged = copy.deepcopy(base_orders)
    for farm, ports in carry.items():
        merged.setdefault(farm, {})
        for port, data in ports.items():
            merged[farm].setdefault(port, {'pallets': 0, 'cajas': 0, 'pallets_by_size': {}})
            merged[farm][port]['pallets'] += data.get('pallets', 0)
            merged[farm][port]['cajas']   += data.get('cajas', 0)
    return merged


def _compute_original_plan(orders, sorted_days,
                            unavailable_vehicle_ids_by_day=None,
                            min_dispatch=None):
    """
    Plan original día a día con carry-forward de residuos pequeños.
    Viajes solo con < min_dispatch pallets no se despachan ese día —
    sus pallets pasan al día siguiente.
    Retorna {day: [export_trips]}.
    """
    if unavailable_vehicle_ids_by_day is None:
        unavailable_vehicle_ids_by_day = {}
    if min_dispatch is None:
        min_dispatch = _MIN_DISPATCH_ORIG

    day_trips = {}
    carry     = {}

    for idx, day in enumerate(sorted_days):
        is_last = (idx == len(sorted_days) - 1)
        today   = _merge_carry(orders.get(day, {}), carry)
        carry   = {}

        unavail   = unavailable_vehicle_ids_by_day.get(day, set())
        all_trips = optimize_day(today, unavailable_vehicle_ids=unavail, relaxed=True, cap_mediodia=True)
        exp_trips = [t for t in all_trips if t.get('trip_type') == 'export']

        final = []
        for t in exp_trips:
            pallets = t.get('pallets_cargados', 0)
            farms   = t.get('farms', {})
            # Viaje solo (1 finca) con pallets menores al umbral → diferir
            if not is_last and len(farms) == 1 and 0 < pallets < min_dispatch:
                farm = next(iter(farms))
                fd   = farms[farm]
                port = next(iter(today.get(farm, {}).keys()), 'PUERTO ANTIOQUIA')
                carry.setdefault(farm, {}).setdefault(port, {'pallets': 0, 'cajas': 0})
                carry[farm][port]['pallets'] += pallets
                carry[farm][port]['cajas']   += fd.get('cajas', 0)
            else:
                final.append(t)

        if final:
            day_trips[day] = final

    return day_trips


def write_suggested_pedido_sheet(wb, orders_orig, adjusted_orders, moves,
                                 semana_num, unavailable_vehicle_ids_by_day=None,
                                 sheet_name='PLAN DE DESPACHO',
                                 sheet_title=None,
                                 relaxed=False,
                                 precomputed_trips=None):
    """Hoja PLAN DE DESPACHO o PLAN DESPACHO ORIGINAL: resumen ejecutivo."""
    if unavailable_vehicle_ids_by_day is None:
        unavailable_vehicle_ids_by_day = {}

    if sheet_title is None:
        sheet_title = sheet_name
    ws   = wb.create_sheet(sheet_name)
    dias = [d for d in DAY_ORDER if d in orders_orig]

    C_TITLE  = '1B5E20'
    C_ORIG   = '1565C0'
    C_OPT    = '2E7D32'
    C_AHORRO = 'BF360C'
    C_MOVE   = 'E65100'
    ALT      = ['FFFFFF', 'F8F9FA']

    # Detectar si hay ajustes reales entre días
    def farm_totals(day_ord):
        return {fn: sum(d.get('pallets',0) for d in ports.values())
                for fn, ports in day_ord.items()}

    has_any_adjustment = any(
        farm_totals(orders_orig.get(d,{})) != farm_totals(adjusted_orders.get(d,{}))
        for d in dias
    )

    NCOLS = 10
    ws.column_dimensions['A'].width = 14
    for ci in range(2, 6):
        ws.column_dimensions[get_column_letter(ci)].width = 11
    for ci in range(6, 10):
        ws.column_dimensions[get_column_letter(ci)].width = 11
    ws.column_dimensions['J'].width = 14

    def cf(c, value=None, bold=False, bg=None, color='000000',
           halign='center', size=9, num_fmt=None, italic=False):
        if value is not None:
            c.value = value
        c.font      = Font(name='Arial', size=size, bold=bold, color=color, italic=italic)
        if bg:
            c.fill  = PatternFill('solid', fgColor=bg)
        c.alignment = Alignment(horizontal=halign, vertical='center', wrap_text=True)
        if num_fmt:
            c.number_format = num_fmt

    row = 1

    # ── Título ───────────────────────────────────────────────────────────────
    ws.merge_cells(start_row=row, start_column=1, end_row=row, end_column=NCOLS)
    cf(ws.cell(row,1), '{} — SEMANA {}'.format(sheet_title, semana_num),
       bold=True, bg=C_TITLE, color='FFFFFF', size=13)
    ws.row_dimensions[row].height = 28
    row += 2

    if has_any_adjustment:
        # ════════════════════════════════════════════════════════════════════
        # CASO: Hay ajustes entre días → mostrar comparación
        # ════════════════════════════════════════════════════════════════════
        ws.merge_cells(start_row=row, start_column=1, end_row=row, end_column=NCOLS)
        cf(ws.cell(row,1),
           '⚠️  Este plan propone ajustes entre días para mejorar la eficiencia. '
           'Revise los cambios y apruebe antes de despachar.',
           bg='FFF3E0', color='BF360C', size=9, bold=True)
        ws.row_dimensions[row].height = 18
        row += 2

        # Cabeceras grupo
        cf(ws.cell(row,1), '', bg='ECEFF1')
        ws.merge_cells(start_row=row, start_column=2, end_row=row, end_column=5)
        cf(ws.cell(row,2), 'PEDIDO ORIGINAL', bold=True, bg=C_ORIG, color='FFFFFF', size=9)
        ws.merge_cells(start_row=row, start_column=6, end_row=row, end_column=9)
        cf(ws.cell(row,6), 'CON AJUSTES', bold=True, bg=C_OPT, color='FFFFFF', size=9)
        cf(ws.cell(row,10), 'AHORRO', bold=True, bg=C_AHORRO, color='FFFFFF', size=9)
        ws.row_dimensions[row].height = 16
        row += 1

        for ci, h in enumerate(['DÍA','Viajes','Cajas','Pallets','Costo',
                                 'Viajes','Cajas','Pallets','Costo','$ Ahorro'], 1):
            cf(ws.cell(row,ci), h, bold=True, bg='ECEFF1', size=8)
        border_all(ws, row-1, row, 1, NCOLS)
        ws.row_dimensions[row].height = 14
        row += 1

        grand_o = {'v':0,'c':0,'p':0,'co':0}
        grand_a = {'v':0,'c':0,'p':0,'co':0}
        for ri, dia in enumerate(dias):
            unavail = unavailable_vehicle_ids_by_day.get(dia, set())
            o_exp = [t for t in optimize_day(orders_orig.get(dia,{}),
                     unavailable_vehicle_ids=unavail, relaxed=True)
                     if t.get('trip_type')=='export']
            a_ord = adjusted_orders.get(dia,{})
            a_exp = ([t for t in optimize_day(a_ord, unavailable_vehicle_ids=unavail)
                      if t.get('trip_type')=='export'] if a_ord else [])
            ov=len(o_exp); oc=int(sum(sum(f['cajas'] for f in t['farms'].values()) for t in o_exp))
            op=int(sum(t['pallets_cargados'] for t in o_exp)); oco=int(sum(t['costo'] for t in o_exp))
            av=len(a_exp); ac=int(sum(sum(f['cajas'] for f in t['farms'].values()) for t in a_exp))
            ap=int(sum(t['pallets_cargados'] for t in a_exp)); aco=int(sum(t['costo'] for t in a_exp))
            ahorro = oco - aco
            # Destacar filas con ajuste real
            ft_o = farm_totals(orders_orig.get(dia,{}))
            ft_a = farm_totals(adjusted_orders.get(dia,{}))
            adj_bg = 'FFF9C4' if ft_o != ft_a else ALT[ri%2]
            a_col = 'C62828' if ahorro<0 else ('1B5E20' if ahorro>0 else '555555')
            for ci, val in enumerate([dia.capitalize(), ov, oc, op, oco,
                                       av, ac, ap, aco,
                                       ahorro if ahorro!=0 else '-'], 1):
                c = ws.cell(row,ci)
                nf = ('"$"#,##0' if ci in (5,9) else ('#,##0' if ci in (3,7) else None))
                if ci==10:
                    cf(c, ahorro if ahorro!=0 else '-', size=8, bg=adj_bg,
                       num_fmt='"$"#,##0' if ahorro!=0 else None,
                       color=a_col, bold=(ahorro!=0))
                else:
                    cf(c, val, size=8, bg=adj_bg, num_fmt=nf or '')
            border_all(ws, row, row, 1, NCOLS)
            ws.row_dimensions[row].height = 14
            grand_o['v']+=ov; grand_o['c']+=oc; grand_o['p']+=op; grand_o['co']+=oco
            grand_a['v']+=av; grand_a['c']+=ac; grand_a['p']+=ap; grand_a['co']+=aco
            row += 1

        tot_ah = grand_o['co'] - grand_a['co']
        a_col  = '1B5E20' if tot_ah>0 else ('C62828' if tot_ah<0 else '555555')
        for ci, (val,fmt) in enumerate([
            ('TOTAL SEMANA',None),(grand_o['v'],None),(grand_o['c'],'#,##0'),
            (grand_o['p'],None),(grand_o['co'],'"$"#,##0'),
            (grand_a['v'],None),(grand_a['c'],'#,##0'),
            (grand_a['p'],None),(grand_a['co'],'"$"#,##0'),
            (tot_ah,'"$"#,##0')], 1):
            cf(ws.cell(row,ci), val, bold=True, bg='E0E0E0', size=9,
               num_fmt=fmt or '', color=a_col if ci==10 else '000000')
        border_all(ws, row, row, 1, NCOLS)
        ws.row_dimensions[row].height = 18
        row += 2

        # Ajustes detallados
        ws.merge_cells(start_row=row, start_column=1, end_row=row, end_column=6)
        cf(ws.cell(row,1), 'AJUSTES PROPUESTOS ENTRE DÍAS',
           bold=True, bg=C_MOVE, color='FFFFFF', size=10)
        ws.row_dimensions[row].height = 20
        row += 1
        for ci, h in enumerate(['Tipo','Finca','De','A','Pallets','Motivo'], 1):
            cf(ws.cell(row,ci), h, bold=True, bg='FFE0B2', color='BF360C', size=9)
        border_all(ws, row, row, 1, 6)
        ws.row_dimensions[row].height = 14
        row += 1
        for m in moves:
            tipo = 'Diferimiento' if m['type']=='deferral' else 'Anticipación'
            bg2  = 'FFF3E0' if m['type']=='deferral' else 'E8F5E9'
            cf(ws.cell(row,1), tipo,                   bg=bg2, size=9)
            cf(ws.cell(row,2), m['farm'].title(),       bg=bg2, size=9, halign='left')
            cf(ws.cell(row,3), m['from_day'].title(),   bg=bg2, size=9)
            cf(ws.cell(row,4), m['to_day'].title(),     bg=bg2, size=9)
            cf(ws.cell(row,5), m['pallets'],            bg=bg2, size=9)
            cf(ws.cell(row,6), m.get('reason',''),      bg=bg2, size=9, halign='left')
            ws.column_dimensions['F'].width = 45
            border_all(ws, row, row, 1, 6)
            ws.row_dimensions[row].height = 14
            row += 1
        row += 1

    else:
        # ════════════════════════════════════════════════════════════════════
        # CASO: Sin ajustes → mostrar solo el plan óptimo
        # ════════════════════════════════════════════════════════════════════
        ws.merge_cells(start_row=row, start_column=1, end_row=row, end_column=NCOLS)
        cf(ws.cell(row,1),
           '✅  Plan optimizado respetando exactamente las fechas del pedido — sin necesidad de ajustes entre días.',
           bg='E8F5E9', color='1B5E20', size=9, bold=True)
        ws.row_dimensions[row].height = 18
        row += 2

        # Tabla resumen simple (solo plan óptimo)
        ws.merge_cells(start_row=row, start_column=1, end_row=row, end_column=6)
        cf(ws.cell(row,1), 'RESUMEN DE LA SEMANA',
           bold=True, bg='263238', color='FFFFFF', size=10)
        ws.row_dimensions[row].height = 20
        row += 1

        for ci, h in enumerate(['DÍA','Viajes','Cajas','Pallets','Costo Total',''], 1):
            cf(ws.cell(row,ci), h, bold=True, bg='ECEFF1', size=8)
        border_all(ws, row, row, 1, 5)
        ws.row_dimensions[row].height = 14
        row += 1

        grand = {'v':0,'c':0,'p':0,'co':0}
        for ri, dia in enumerate(dias):
            unavail = unavailable_vehicle_ids_by_day.get(dia,set())
            a_ord   = adjusted_orders.get(dia,{})
            if precomputed_trips is not None:
                a_exp = precomputed_trips.get(dia, [])
            else:
                a_exp = ([t for t in optimize_day(a_ord, unavailable_vehicle_ids=unavail, relaxed=relaxed)
                          if t.get('trip_type')=='export'] if a_ord else [])
            av=len(a_exp); ac=int(sum(sum(f['cajas'] for f in t['farms'].values()) for t in a_exp))
            ap=int(sum(t['pallets_cargados'] for t in a_exp)); aco=int(sum(t['costo'] for t in a_exp))
            bg = ALT[ri%2]
            for ci, (val,fmt) in enumerate([
                (dia.capitalize(),None),(av,None),(ac,'#,##0'),
                (ap,None),(aco,'"$"#,##0'),(None,None)], 1):
                if val is not None:
                    cf(ws.cell(row,ci), val, size=9, bg=bg, num_fmt=fmt or '')
            border_all(ws, row, row, 1, 5)
            ws.row_dimensions[row].height = 14
            grand['v']+=av; grand['c']+=ac; grand['p']+=ap; grand['co']+=aco
            row += 1

        for ci, (val,fmt) in enumerate([
            ('TOTAL SEMANA',None),(grand['v'],None),(grand['c'],'#,##0'),
            (grand['p'],None),(grand['co'],'"$"#,##0'),(None,None)], 1):
            if val is not None:
                cf(ws.cell(row,ci), val, bold=True, bg='1B5E20',
                   color='FFFFFF', size=9, num_fmt=fmt or '')
        border_all(ws, row, row, 1, 5)
        ws.row_dimensions[row].height = 18
        row += 2

    # ════════════════════════════════════════════════════════════════════════
    # DETALLE DE VIAJES — plan óptimo por día
    # ════════════════════════════════════════════════════════════════════════
    ws.merge_cells(start_row=row, start_column=1, end_row=row, end_column=NCOLS)
    cf(ws.cell(row,1), 'DETALLE DE VIAJES POR DÍA',
       bold=True, bg='263238', color='FFFFFF', size=10)
    ws.row_dimensions[row].height = 20
    row += 1

    for dia in dias:
        unavail = unavailable_vehicle_ids_by_day.get(dia,set())
        a_ord   = adjusted_orders.get(dia,{})
        if precomputed_trips is not None:
            a_trips = precomputed_trips.get(dia, [])
        else:
            a_trips = ([t for t in optimize_day(a_ord, unavailable_vehicle_ids=unavail, relaxed=relaxed)
                        if t.get('trip_type')=='export'] if a_ord else [])

        ws.merge_cells(start_row=row, start_column=1, end_row=row, end_column=NCOLS)
        cf(ws.cell(row,1),
           '{} {}  —  {} viajes  ·  {}P  ·  {:,} cajas  ·  ${:,}'.format(
               DAY_EMOJIS.get(dia,''), dia.title(),
               len(a_trips),
               sum(t.get('pallets_cargados',0) for t in a_trips),
               int(sum(sum(f.get('cajas',0) for f in t['farms'].values()) for t in a_trips)),
               int(sum(t.get('costo',0) for t in a_trips))),
           bold=True, bg=DAY_COLORS.get(dia, C_TITLE), color='FFFFFF', size=10)
        ws.row_dimensions[row].height = 18
        row += 1

        for ci, h in enumerate(['Conductor','Hora',
                                 'Distribución por finca (pallets)',
                                 '','','','','','Pallets','Costo'], 1):
            if h:
                cf(ws.cell(row,ci), h, bold=True, bg='ECEFF1', size=8)
        ws.merge_cells(start_row=row, start_column=3, end_row=row, end_column=8)
        border_all(ws, row, row, 1, NCOLS)
        ws.row_dimensions[row].height = 14
        row += 1

        for ti, t in enumerate(a_trips):
            bg2  = ALT[ti%2]
            fstr = '  ·  '.join('{}: {}P'.format(fn.title(), fd.get('pallets',0))
                                 for fn, fd in t.get('farms',{}).items())
            cf(ws.cell(row,1), t.get('conductor',''), halign='left', size=8, bg=bg2)
            cf(ws.cell(row,2), t.get('hora',''),      size=8, bg=bg2)
            ws.merge_cells(start_row=row, start_column=3, end_row=row, end_column=8)
            cf(ws.cell(row,3), fstr, halign='left', size=8, bg=bg2)
            cf(ws.cell(row,9),  t.get('pallets_cargados',0), size=8, bg=bg2)
            cf(ws.cell(row,10), t.get('costo',0), size=8, bg=bg2, num_fmt='"$"#,##0')
            border_all(ws, row, row, 1, NCOLS)
            ws.row_dimensions[row].height = 16
            row += 1

        # Fila subtotal del día
        day_p  = sum(t.get('pallets_cargados',0) for t in a_trips)
        day_c  = int(sum(sum(f.get('cajas',0) for f in t['farms'].values()) for t in a_trips))
        day_co = int(sum(t.get('costo',0) for t in a_trips))
        day_color = DAY_COLORS.get(dia, C_TITLE)
        ws.merge_cells(start_row=row, start_column=1, end_row=row, end_column=7)
        cf(ws.cell(row,1),
           'SUBTOTAL {} — {} viajes  ·  {:,} cajas'.format(
               dia.upper(), len(a_trips), day_c),
           bold=True, bg=day_color, color='FFFFFF', size=9)
        cf(ws.cell(row,9),  int(day_p),  bold=True, bg=day_color, color='FFFFFF', size=9)
        cf(ws.cell(row,10), day_co, bold=True, bg=day_color, color='FFFFFF', size=9,
           num_fmt='"$"#,##0')
        border_all(ws, row, row, 1, NCOLS)
        ws.row_dimensions[row].height = 16
        row += 2

    ws.freeze_panes = 'B4'


def border_all(ws, min_row, max_row, min_col, max_col):
    thin = Side(style='thin', color='CCCCCC')
    b    = Border(left=thin, right=thin, top=thin, bottom=thin)
    for r in range(min_row, max_row + 1):
        for c in range(min_col, max_col + 1):
            ws.cell(r, c).border = b


def _write_trip_section(ws, start_row, trips, color, title, NCOLS):
    """Escribe una sección de viajes (título + cabeceras + filas + total).
    Retorna (next_row, total_cajas, total_pallets, total_costo, n_viajes)."""
    row = start_row
    alt = ['FFFFFF', 'F1F8E9']

    # Cabecera de sección
    ws.merge_cells(start_row=row, start_column=1, end_row=row, end_column=NCOLS)
    c = ws.cell(row, 1, title)
    c.font      = Font(name='Arial', size=10, bold=True, color='FFFFFF')
    c.fill      = PatternFill('solid', fgColor=color)
    c.alignment = Alignment(horizontal='center', vertical='center')
    ws.row_dimensions[row].height = 20
    row += 1

    # Cabeceras columnas
    headers    = ['#', 'CONDUCTOR', 'VEHICULO', 'CAP.', 'RUTA',
                  'CAJAS POR FINCA', 'PALLETS POR FINCA',
                  'TOTAL CAJAS', 'PALLETS', 'HORA', 'ENTREGAR EN', 'COSTO']
    for ci, h in enumerate(headers, 1):
        c = ws.cell(row, ci, h)
        c.font      = Font(name='Arial', size=9, bold=True, color='FFFFFF')
        c.fill      = PatternFill('solid', fgColor=color)
        c.alignment = Alignment(horizontal='center', vertical='center', wrap_text=True)
    ws.row_dimensions[row].height = 22
    row += 1

    sec_cajas = 0; sec_pallets = 0; sec_costo = 0; sec_viajes = 0
    trip_num = 0

    for trip in trips:
        if trip.get('trip_type') != 'export':
            continue
        trip_num  += 1
        hora        = trip.get('hora', '')
        conductor   = trip.get('conductor', '')
        vehicle     = VEHICLE_DISPLAY.get(trip.get('vehicle_id', ''), {}).get(
                          'vehicle', trip.get('vehicle_id', ''))
        cap         = trip.get('capacidad', '')
        zone        = trip.get('zone', '')
        farms       = trip.get('farms', {})
        pallets     = trip.get('pallets_cargados', 0)
        costo       = trip.get('costo', 0)
        destination = trip.get('destination', '')
        farm_psize      = trip.get('farm_pallets_size', {})
        farm_total_pals = trip.get('farm_total_pallets', {})

        farm_zones_in_trip = {FARM_ZONES.get(f, zone) for f in farms}
        if len(farm_zones_in_trip) > 1:
            zone_label = 'Chigorodo -> Apartado'
        else:
            zone_label = list(farm_zones_in_trip)[0].title() if farm_zones_in_trip else zone.title()
        if len(farms) == 1:
            fname = list(farms.keys())[0]
            ruta  = '-> {} ({})'.format(fname.title(), zone_label)
        else:
            ruta  = '-> {} ({})'.format(' -> '.join(f.title() for f in farms), zone_label)

        cajas_total = sum(d.get('cajas', 0) for d in farms.values())
        cajas_str   = '  |  '.join(
            '{}: {:,}'.format(fn.title(), fd.get('cajas', 0))
            for fn, fd in farms.items())
        pallet_parts = []
        for fn, fd in farms.items():
            psz = farm_psize.get(fn, {})
            assigned_p = int(fd.get('pallets', 0))
            total_p    = int(farm_total_pals.get(fn, assigned_p) or assigned_p or 1)
            lbl = pallet_size_label(assigned_p, total_p, psz) \
                  if psz else '{}P'.format(assigned_p)
            pallet_parts.append('{}: {}'.format(fn.title(), lbl))
        pallets_str = '  |  '.join(pallet_parts)

        bg   = alt[(row - start_row) % 2]
        vals = [trip_num, conductor, vehicle, cap, ruta,
                cajas_str, pallets_str,
                int(cajas_total), int(pallets), hora, destination, costo]
        for ci, val in enumerate(vals, 1):
            c = ws.cell(row, ci, val)
            c.font      = Font(name='Arial', size=9)
            c.fill      = PatternFill('solid', fgColor=bg)
            c.alignment = Alignment(
                horizontal='left' if ci in (5, 6, 7) else 'center',
                vertical='center', wrap_text=True)
            if ci == 12: c.number_format = '"$"#,##0'
            if ci == 8:  c.number_format = '#,##0'
        ws.row_dimensions[row].height = 22
        sec_cajas   += cajas_total
        sec_pallets += pallets
        sec_costo   += costo
        sec_viajes  += 1
        row += 1

    # Fila subtotal
    if sec_viajes > 0:
        ws.merge_cells(start_row=row, start_column=1, end_row=row, end_column=7)
        c = ws.cell(row, 1, 'SUBTOTAL — {} viajes'.format(sec_viajes))
        c.font      = Font(name='Arial', size=9, bold=True, color='FFFFFF')
        c.fill      = PatternFill('solid', fgColor=color)
        c.alignment = Alignment(horizontal='center', vertical='center')
        for ci, (val, fmt) in enumerate([(int(sec_cajas), '#,##0'),
                                          (int(sec_pallets), '#,##0'),
                                          ('', None), ('', None),
                                          (sec_costo, '"$"#,##0')], 8):
            c = ws.cell(row, ci, val)
            c.font      = Font(name='Arial', size=9, bold=True, color='FFFFFF')
            c.fill      = PatternFill('solid', fgColor=color)
            c.alignment = Alignment(horizontal='center', vertical='center')
            if fmt: c.number_format = fmt
        ws.row_dimensions[row].height = 18
        border_all(ws, start_row + 2, row, 1, NCOLS)
        row += 1
    else:
        ws.merge_cells(start_row=row, start_column=1, end_row=row, end_column=NCOLS)
        c = ws.cell(row, 1, 'Sin viajes este día')
        c.font      = Font(name='Arial', size=9, italic=True, color='888888')
        c.fill      = PatternFill('solid', fgColor='F5F5F5')
        c.alignment = Alignment(horizontal='center', vertical='center')
        ws.row_dimensions[row].height = 15
        row += 1

    return row, sec_cajas, sec_pallets, sec_costo, sec_viajes


def write_day_sheet(wb, day, orig_trips, sug_trips, first_sheet, banafrut_orders=None, inter_day_moves=None, has_modifications=False):
    NCOLS   = 12
    color   = DAY_COLORS.get(day, '1B5E20')
    ORIG_C  = '1565C0'   # azul — Banafrut original
    SUG_C   = '2E7D32'   # verde — sugerido

    if first_sheet:
        ws = wb.active; ws.title = day
    else:
        ws = wb.create_sheet(title=day)

    # Anchos de columna
    col_widths = [4, 16, 22, 6, 38, 30, 28, 11, 9, 10, 16, 14]
    for ci, w in enumerate(col_widths, 1):
        ws.column_dimensions[get_column_letter(ci)].width = w

    # Calcular totales sugeridos para el título
    sug_exp = [t for t in sug_trips if t.get('trip_type') == 'export']
    sug_c   = sum(sum(f.get('cajas',0) for f in t['farms'].values()) for t in sug_exp)
    sug_p   = sum(t.get('pallets_cargados', 0) for t in sug_exp)
    sug_co  = sum(t.get('costo', 0) for t in sug_exp)

    # Fila 1: título
    ws.merge_cells(start_row=1, start_column=1, end_row=1, end_column=NCOLS)
    c = ws.cell(1, 1, '{} {}  —  {:,} cajas  ·  {}P  ·  ${:,}'.format(
        DAY_EMOJIS.get(day, ''), day, int(sug_c), int(sug_p), int(sug_co)))
    c.font      = Font(name='Arial', size=12, bold=True, color='FFFFFF')
    c.fill      = PatternFill('solid', fgColor=color)
    c.alignment = Alignment(horizontal='center', vertical='center')
    ws.row_dimensions[1].height = 28

    row = 2

    if not has_modifications:
        # ── Sin cambios al pedido de Banafrut: mostrar solo la ruta óptima ──
        _title = 'DESPACHO ÓPTIMO  ({}P  ·  {:,} cajas  ·  ${:,})'.format(
            int(sug_p), int(sug_c), int(sug_co))
        row, s_cajas, s_pallets, s_costo, s_viajes = _write_trip_section(
            ws, row, sug_trips, SUG_C, _title, NCOLS)
    else:
        # ── Pedido modificado: mostrar Banafrut original arriba, sugerencia abajo ──
        # Sección 1: PEDIDO BANAFRUT
        b_ordered_p = int(sum(d.get('pallets', 0) for d in (banafrut_orders or {}).values()))
        b_ordered_c = int(sum(d.get('cajas', 0)   for d in (banafrut_orders or {}).values()))
        b_orig_exp  = [t for t in orig_trips if t.get('trip_type') == 'export']
        b_shipped_p = int(sum(t.get('pallets_cargados', 0) for t in b_orig_exp))
        b_shipped_c = int(sum(sum(f.get('cajas', 0) for f in t['farms'].values()) for t in b_orig_exp))
        b_unshipped = max(0, b_ordered_p - b_shipped_p)
        if b_unshipped > 0:
            _b_title = 'PEDIDO BANAFRUT — Sin modificaciones  ({} ordenados | {} enviados | ⚠️ {}P sin despachar)'.format(
                '{}P · {:,}c'.format(b_ordered_p, b_ordered_c),
                '{}P · {:,}c'.format(b_shipped_p, b_shipped_c),
                b_unshipped)
        else:
            _b_title = 'PEDIDO BANAFRUT — Sin modificaciones  ({}P  ·  {:,} cajas  ·  ${:,})'.format(
                b_shipped_p, b_shipped_c,
                int(sum(t.get('costo', 0) for t in b_orig_exp)))
        row, b_cajas, b_pallets, b_costo, b_viajes = _write_trip_section(
            ws, row, orig_trips, ORIG_C, _b_title, NCOLS)
        row += 1  # espacio

        # Sección 2: NUESTRA SUGERENCIA (con movimientos entre días)
        _moves = inter_day_moves or {}
        _defer_out = sum(m.get('pallets', 0) for m in _moves.get(day, []) if m.get('type') == 'deferral')
        _moved_to  = sum(
            m.get('pallets', 0)
            for d2, mvs in _moves.items()
            for m in mvs
            if m.get('type') == 'deferral' and m.get('to_day') == day
        )
        _extra = []
        if _defer_out:
            # Obtener nombre del día destino (puede ser anterior o siguiente)
            _dest_days = list({m.get('to_day','').title()
                               for m in _moves.get(day, []) if m.get('type') == 'deferral'})
            _dest_str = ' / '.join(_dest_days) if _dest_days else 'otro día'
            _extra.append('↓ {}P trasladados al {}'.format(_defer_out, _dest_str))
        if _moved_to:
            # Obtener nombre del día origen
            _orig_days = list({m.get('from_day','').title()
                               for d2, mvs in _moves.items()
                               for m in mvs
                               if m.get('type') == 'deferral' and m.get('to_day') == day})
            _orig_str = ' / '.join(_orig_days) if _orig_days else 'otro día'
            _extra.append('↑ {}P recibidos del {}'.format(_moved_to, _orig_str))
        _move_str = '  |  '.join(_extra)
        _s_title = 'NUESTRA SUGERENCIA — Optimizado  ({}P  ·  {:,} cajas  ·  ${:,})  {}'.format(
            int(sug_p), int(sug_c), int(sug_co),
            '  ⚠️ ' + _move_str if _move_str else '')
        row, s_cajas, s_pallets, s_costo, s_viajes = _write_trip_section(
            ws, row, sug_trips, SUG_C, _s_title, NCOLS)

    ws.freeze_panes = 'A3'
    return s_costo, s_cajas, s_pallets, s_viajes

def generate_excel_bytes(orders, semana_num, unavailable_vehicle_ids_by_day=None):
    if unavailable_vehicle_ids_by_day is None:
        unavailable_vehicle_ids_by_day = {}
    wb          = Workbook()
    sorted_days = sorted(orders.keys(),
                         key=lambda d: DAY_ORDER.index(d) if d in DAY_ORDER else 99)
    # Plan original con carry-forward (reemplaza el loop relaxed=True directo)
    # Se calcula una sola vez y se reutiliza en todas las hojas
    orig_plan_trips = _compute_original_plan(
        orders, sorted_days, unavailable_vehicle_ids_by_day
    )
    day_results = orig_plan_trips   # alias para compatibilidad con código posterior
    if not day_results:
        return None, {}

    # Plan sugerido (pedido ajustado con consolidaciones inter-día)
    adjusted_orders_gen, _ = compute_inter_day_moves(orders)
    sug_day_results = {}
    for day in sorted_days:
        unavail   = unavailable_vehicle_ids_by_day.get(day, set())
        sug_trips = optimize_day(adjusted_orders_gen.get(day, {}),
                                 unavailable_vehicle_ids=unavail)
        if sug_trips:
            sug_day_results[day] = sug_trips

    # Totales generales (sin hoja RESUMEN SEMANA)
    grand = {'cost': 0, 'cajas': 0, 'pallets': 0, 'viajes': 0}
    for day in sorted_days:
        if day not in day_results and day not in sug_day_results:
            continue
        trips        = sug_day_results.get(day, day_results.get(day, []))
        export_trips = [t for t in trips if t.get('trip_type') == 'export']
        grand['cost']    += sum(t['costo'] for t in export_trips)
        grand['cajas']   += sum(sum(f['cajas'] for f in t['farms'].values()) for t in export_trips)
        grand['pallets'] += sum(t['pallets_cargados'] for t in export_trips)
        grand['viajes']  += len(export_trips)

    adjusted_orders_plan, inter_day_moves_plan = compute_inter_day_moves(orders)
    write_suggested_pedido_sheet(
        wb, orders, adjusted_orders_plan, inter_day_moves_plan,
        semana_num, unavailable_vehicle_ids_by_day,
        sheet_name='PLAN DE DESPACHO',
    )
    # Eliminar la hoja vacía por defecto que crea openpyxl
    default_name = wb.sheetnames[0]
    if default_name != 'PLAN DE DESPACHO':
        del wb[default_name]

    # PLAN DESPACHO ORIGINAL — orig_plan_trips ya calculado arriba
    write_suggested_pedido_sheet(
        wb, orders, orders, [],
        semana_num, unavailable_vehicle_ids_by_day,
        sheet_name='PLAN DESPACHO ORIGINAL',
        sheet_title='PLAN DESPACHO ORIGINAL',
        relaxed=True,
        precomputed_trips=orig_plan_trips,
    )

    # Build per-day movement index for write_day_sheet headers
    _raw_moves = inter_day_moves_plan
    inter_day_moves_by_day = {}
    for mv in (_raw_moves or []):
        fd = mv.get('from_day', '')
        td = mv.get('to_day', '')
        inter_day_moves_by_day.setdefault(fd, []).append({**mv, 'type': 'deferral'})
        inter_day_moves_by_day.setdefault(td, []).append({**mv, 'type': 'anticipation'})

    # Determinar qué días tuvieron modificaciones en el pedido (movimientos inter-día)
    def _farm_totals(day_orders):
        return {fn: sum(d.get('pallets', 0) for d in ports.values())
                for fn, ports in day_orders.items()}

    for day in sorted_days:
        if day not in day_results and day not in sug_day_results:
            continue
        orig_t = day_results.get(day, [])
        sug_t  = sug_day_results.get(day, [])
        orig_fp = _farm_totals(orders.get(day, {}))
        sug_fp  = _farm_totals(adjusted_orders_gen.get(day, {}))
        has_mod = (orig_fp != sug_fp)
        write_day_sheet(wb, day, orig_t, sug_t, False,
                        banafrut_orders=orders.get(day, {}),
                        inter_day_moves=inter_day_moves_by_day,
                        has_modifications=has_mod)

    buf = io.BytesIO()
    wb.save(buf)
    return buf.getvalue(), grand
